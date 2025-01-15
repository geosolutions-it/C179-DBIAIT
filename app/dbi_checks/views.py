import os
from django.shortcuts import render
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from django.views.generic.edit import FormView
from django.urls import reverse_lazy
from django.contrib import messages
from app.dbi_checks.forms import ExcelUploadForm
from app.dbi_checks.tasks import copy_to_dbi_files
from app.settings import (
    UPLOADED_XLSX_FILES, 
    YEAR_VALUE,
    DBI_A_1, 
    DBI_A,
    SHEETS_CONFIG,
    EXTRA_SHEETS_CONFIG
)

from openpyxl import load_workbook

# Check: consistenza delle opere
class Consistency_check(LoginRequiredMixin, View):
    def get(self, request):
        return render(request, 'dbi_checks/base-checks.html')

# Check: dati prioritati
class PrioritizedData_check(LoginRequiredMixin, View):
    def get(self, request):
        return render(request, 'dbi_checks/base-checks.html')

class UploadExcelView(LoginRequiredMixin, FormView):
    
    template_name = "dbi_checks/base-checks.html"
    form_class = ExcelUploadForm
    success_url = reverse_lazy("upload-excel-view")

    def form_valid(self, form):
        xlsx_file1 = form.cleaned_data["xlsx_file1"]
        xlsx_file2 = form.cleaned_data["xlsx_file2"]

        xlsx_file1_path = os.path.join(UPLOADED_XLSX_FILES, xlsx_file1.name)
        xlsx_file2_path = os.path.join(UPLOADED_XLSX_FILES, xlsx_file2.name)


        with open(xlsx_file1_path, "wb+") as destination:
            for chunk in xlsx_file1.chunks():
                destination.write(chunk)
        with open(xlsx_file2_path, "wb+") as destination:
            for chunk in xlsx_file2.chunks():
                destination.write(chunk)

        # Get the year from each file
        precending_year = self.get_year(xlsx_file1_path)
        current_year = self.get_year(xlsx_file2_path)
  
        # A temp check, this could be changed
        if precending_year and current_year:
            messages.success(self.request, "Files uploaded and processed successfully!")
            
            # Trigger copy as an async task
            copy_to_dbi_files.send(xlsx_file1_path, DBI_A_1, SHEETS_CONFIG)
            copy_to_dbi_files.send(xlsx_file2_path, DBI_A, {**SHEETS_CONFIG, **EXTRA_SHEETS_CONFIG})

        else:
            messages.error(self.request, "File processing failed. Please check the file content.")

        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Something went wrong with the upload... Please try again")
        return super().form_invalid(form)
    
    def get_year(self, file_path):
        """
        This method get the year from the cell B8 from each
        uploaded xlsx file
        """
         
        try:
            wb = load_workbook(file_path)
            # Get the required sheet and year value
            dati_sheet = wb[YEAR_VALUE["sheet"]]
            year_value = dati_sheet.cell(row=YEAR_VALUE["row"], column=YEAR_VALUE["column"]).value
    
            return year_value

        except Exception as e:
            print(f"Error processing files: {e}")
            return False
