import os
import shutil
import dramatiq
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.formula.translate import Translator
from app.settings import FOR_DOWNLOAD
from app.dbi_checks.utils import get_year, get_last_data_row

import logging

logger = logging.getLogger(__name__)

# Tasks for check: consistenza delle opere 
@dramatiq.actor
def copy_to_dbi_files(file_path, seed, config, formulas_config, file_dependency=None, next_args=None):
    try:

        logger.info(f"Task started with file: {file_path}")

        if file_dependency == True:
            # Create the INPUT.xlsx file which it is needed by the DBI_A formulas
            get_year(file_path)
            try:
                load_workbook(os.path.join(FOR_DOWNLOAD, "INPUT.xlsx"))
            except:
                logger.warning(f"Error: The file INPUT.xlsx did not created !")

        # It's crucial to use the read_only argument because it's quite faster
        up_file = load_workbook(file_path, read_only=True)
        seed_basename = os.path.basename(seed)
        seed_copy = shutil.copy(seed, f"{FOR_DOWNLOAD}/{seed_basename}")
        seed_wb = load_workbook(seed_copy, data_only=False)

        # Iterate over the sheets to copy data
        for source_sheet, config in config.items():
            target_sheet = config["target"]
            
            # Convert column letters to numbers
            min_col = column_index_from_string(config["start_col"])
            min_row = config["start_row"]

            if source_sheet in up_file.sheetnames and target_sheet in seed_wb.sheetnames:

                source = up_file[source_sheet]
                target = seed_wb[target_sheet]

                # Copy data based on the specified column range
                # Usage of chunks to optimize large row ranges
                for row in source.iter_rows(min_row=min_row, max_row=source.max_row, min_col=min_col, max_col=source.max_column):
                    for cell in row:
                        if cell.value is not None:
                            target.cell(row=cell.row, column=cell.column, value=cell.value)

                logger.info(f"Copied data from sheet: {source_sheet} to {target_sheet}")
            else:
                logger.warning(f"Sheet {source_sheet} or {target_sheet} not found!")
    
        # Iterate through each sheet to drag the formulas
        for sheet_name, f_location in formulas_config.items():
            
            if sheet_name in seed_wb.sheetnames:
                sheet = seed_wb[sheet_name]

                # Get column indexes
                start_col_index = column_index_from_string(f_location["start_col"])
                end_col_index = column_index_from_string(f_location["end_col"])
                start_row = f_location["start_row"]
                # Re-definition of the last row because the copied file is processed
                # without saving yet. We don't want to re-load it for time reasons
                last_row = get_last_data_row(sheet)

                # Copy formulas from row 4 to the rest of the rows
                for col_idx in range(start_col_index, end_col_index + 1):
                    column_letter = get_column_letter(col_idx)
                    # Get the formula in row 4
                    formula = sheet[f"{column_letter}{start_row}"].value
                    if isinstance(formula, str) and formula.startswith("="):
                        # Use the Translator to adjust the formula for each subsequent row
                        for row_idx in range(start_row + 1, last_row + 1):
                            translator = Translator(formula, f"{column_letter}{start_row}")
                            adjusted_formula = translator.translate_formula(f"{column_letter}{row_idx}")
                            # Set the adjusted formula in the target row
                            sheet[f"{column_letter}{row_idx}"].value = adjusted_formula

                logger.info(f"The formulas were populated from sheet: {sheet_name}")
            else:
                logger.warning(f"Something went wrong when filling out the formulas !")
        
        
        # Save the changes to the file
        seed_wb.save(seed_copy)

        # Clean up by deleting the temporary files
        os.remove(file_path)
        
        # delete the file INPUT.xlsx if exists
        if os.path.exists(os.path.join(FOR_DOWNLOAD, "INPUT.xlsx")):
            # Delete the file
            os.remove(os.path.join(FOR_DOWNLOAD, "INPUT.xlsx"))

        # If next_args are provided, trigger the next task
        if next_args:
            copy_to_dbi_files.send(*next_args)
        
    except Exception as e:
        print(f"Error processing files in the background: {e}")