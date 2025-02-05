import os
import logging
from openpyxl import load_workbook, Workbook

from django.conf import settings

logger = logging.getLogger(__name__)

class CheckType:
    CDO = "CDO"
    DP = "DP"

class YearHandler:
    def __init__(self, imported_file, export_dir=None):
        
        self.imported_file = imported_file
        self.export_dir = export_dir

    def get_year(self):
        """
        Retrieve the year from the specified cell in the imported file.

        Returns:
            int or bool: The year value if successful, False if an error occurs.
        """
        try:
            # Open the workbook in read-only mode for performance
            wb1 = load_workbook(self.imported_file, read_only=True, data_only=True)
            
            # Access the sheet and cell specified in the settings
            dati_sheet = wb1[settings.YEAR_VALUE["sheet"]]
            year_value = dati_sheet.cell(
                row=settings.YEAR_VALUE["row"], 
                column=settings.YEAR_VALUE["column"]
            ).value

            return year_value

        except Exception as e:
            print(f"Error retrieving year: {e}")
            return False

    def set_year_to_file(self):
        """
        Create the INPUT.xlsx file with the year value.

        Parameters:
            year_value (int): The year value to write into the INPUT.xlsx file.
        
        Returns:
            bool: True if the file was successfully created, False otherwise.
        """
        
        year_value = self.get_year()
        if not year_value:
            print("Failed to retrieve the year, cannot create INPUT.xlsx.")
            return False
 
        try:
            # Create a new workbook and sheet
            wb2 = Workbook()
            ws = wb2.active
            ws.title = "Input anno"

            # Write the year value to cell A1
            ws['A1'] = year_value

            # Save the workbook to the export directory
            wb2.save(os.path.join(self.export_dir, "INPUT.xlsx"))
            return True

        except Exception as e:
            print(f"Error creating INPUT.xlsx: {e}")
            return False