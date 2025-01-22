import os
import shutil
import pathlib
import traceback

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.formula.translate import Translator

from django.db.models import Q, ObjectDoesNotExist
from django.contrib.auth import get_user_model
from django.utils import timezone

from app.settings import FOR_DOWNLOAD
from app.dbi_checks.utils import get_year, get_last_data_row, TaskType_CheckDbi
from app.dbi_checks.models import Task_CheckDbi, Xlsx, ImportedSheet 
from app.dbi_checks.utils import import_sheet

from app.scheduler.tasks.base_task import BaseTask, trace_it
from app.scheduler.utils import TaskStatus, Schema
from app.scheduler import exceptions
from app.scheduler.logging import Tee

import logging

logger = logging.getLogger(__name__)

class Import_DbiCheckTask(BaseTask):
    """
    Dramatiq Import task definition class.

    Example usage:
        user = User.objects.get(...)
        gpkg_path = pathlib.Path(...)

        try:
            ImportTask.send(ImportTask.pre_send(requesting_user=user, gpkg_path=gpkg_path))
        except scheduler.exceptions.QueuingCriteriaViolated as e:
            logger.error('Scheduling criteria violated for Import task')
    """

    task_type = TaskType_CheckDbi.IMPORT_CheckDbi
    name = "import_check_dbi"
    schema = Schema.ANALYSIS

    @classmethod
    def pre_send(
        cls,
        requesting_user: get_user_model(),
        file_path1: str,
        file_path2: str,
        seed_a: str,
        seed_a_1: str, 
        dbi_a_config: str,
        dbi_a_1_config: str,
        dbi_a_formulas: str,
        dbi_a_1_formulas: str,
        file_dependency=None,
    ):

        # Check if the xlsx files exists
        file1 = pathlib.Path(file_path1)
        file2 = pathlib.Path(file_path2)
        if not file1.exists() or not file2.exists():
            raise exceptions.SchedulingParametersError(
                f"Provided *.xlsx files do not exist: {file1.name}, {file2.name}"
            )

        colliding_tasks = Task_CheckDbi.objects.filter(
            Q(status=TaskStatus.QUEUED) | Q(status=TaskStatus.RUNNING)
        ).exclude(Q(schema=Schema.ANALYSIS) & Q(type=TaskType_CheckDbi.PROCESS_CheckDbi))

        if len(colliding_tasks) > 0:
            raise exceptions.QueuingCriteriaViolated(
                "Qualcosa è andato storto durante il caricamento o l'elaborazione"
                "Si prega di riprovare più tardi"
                # f"Following tasks prevent scheduling this operation: {[task.id for task in colliding_tasks]}"
            )

        # Get or create Xlsx ORM model instance for this task execution
        xlsx, created = Xlsx.objects.get_or_create(name=f"{file1.name.split('.')[0]}_{file2.name.split('.')[0]}",
                                                   file_path1=file_path1,
                                                   file_path2=file_path2
                                                   )
        if created:
            xlsx.save()

        # Create Task ORM model instance for this task execution
        current_task = Task_CheckDbi(
            requesting_user=requesting_user,
            schema=cls.schema,
            xlsx=xlsx,
            type=cls.task_type,
            name=cls.name,
            params={
                "args": [
                  str(file_path1),
                  str(file_path2),
                  seed_a,
                  seed_a_1,
                  dbi_a_config,
                  dbi_a_1_config,
                  dbi_a_formulas,
                  dbi_a_1_formulas

                ],
                "kwargs": {
                    "file_dependency": file_dependency,
                }
            }
        )
        current_task.save()

        return current_task.id

    @trace_it
    def execute(self, 
                task_id: int,
                *args,
                **kwargs
                ) -> None:
        
        try:
            (
            file_path1,
            file_path2,
            seed_a,
            seed_a_1,
            dbi_a_config,
            dbi_a_1_config,
            dbi_a_formulas,
            dbi_a_1_formulas,
            ) = args
            
            file_dependency = kwargs.get("file_dependency")

            if file_dependency is True:
                # Create the INPUT.xlsx file which it is needed by the DBI_A formulas
                get_year(file_path1)
                try:
                    load_workbook(os.path.join(FOR_DOWNLOAD, "INPUT.xlsx"))
                except:
                    logger.warning(f"Error: The file INPUT.xlsx did not created !")

            # Copy the first file using DBI_A seed
            logger.info(f"Task started with file: {file_path1}")
            result = self.copy_files(task_id, file_path1, seed_a, dbi_a_config, dbi_a_formulas)
            
            # Copy the second file using the DBI_A_1 seed
            if result is True:
                logger.info(f"Task started with file: {file_path2}")
                self.copy_files(task_id, file_path2, seed_a_1, dbi_a_1_config, dbi_a_1_formulas)

            return True
        
        except Exception as e:
            print(f"Error processing files in the background: {e}")
        
    def copy_files(self, task, file_path, seed, config, formulas_config):

        # It's crucial to use the read_only argument because it's quite faster
        up_file = load_workbook(file_path, read_only=True)
        seed_basename = os.path.basename(seed)
        seed_copy = shutil.copy(seed, f"{FOR_DOWNLOAD}/{seed_basename}")
        seed_wb = load_workbook(seed_copy, data_only=False)

        # Iterate over the sheets to copy data
        for source_sheet, config in config.items():
            
            start_date = timezone.now()
            
            target_sheet = config["target"]
            
            # Convert column letters to numbers
            min_col = column_index_from_string(config["start_col"])
            min_row = config["start_row"]

            if source_sheet in up_file.sheetnames and target_sheet in seed_wb.sheetnames:

                source = up_file[source_sheet]
                target = seed_wb[target_sheet]

                # Copy data based on the specified column range
                # Usage of chunks to optimize large row ranges
                for row in source.iter_rows(min_row=min_row, max_row=source.max_row, min_col=min_col, max_col=source.max_column):
                    for cell in row:
                        if cell.value is not None:
                            target.cell(row=cell.row, column=cell.column, value=cell.value)

                logger.info(f"Copied data from sheet: {source_sheet} to {target_sheet}")
                
                # Call the import_sheet function with a SUCCESS status
                end_date = timezone.now()
                import_sheet(task, source_sheet, os.path.basename(file_path), start_date, end_date, TaskStatus.SUCCESS)
                    
            else:
                end_date = timezone.now()
                import_sheet(task, source_sheet, os.path.basename(file_path), start_date, end_date, TaskStatus.FAILED)
                logger.warning(f"Sheet {source_sheet} or {target_sheet} not found!")

        # Iterate through each sheet to drag the formulas
        for sheet_name, f_location in formulas_config.items():
            
            if sheet_name in seed_wb.sheetnames:
                sheet = seed_wb[sheet_name]

                # Get column indexes
                start_col_index = column_index_from_string(f_location["start_col"])
                end_col_index = column_index_from_string(f_location["end_col"])
                start_row = f_location["start_row"]
                # Re-definition of the last row because the copied file is processed
                # without saving yet. We don't want to re-load it for time reasons
                last_row = get_last_data_row(sheet)

                # Copy formulas from row 4 to the rest of the rows
                for col_idx in range(start_col_index, end_col_index + 1):
                    column_letter = get_column_letter(col_idx)
                    # Get the formula in row 4
                    formula = sheet[f"{column_letter}{start_row}"].value
                    if isinstance(formula, str) and formula.startswith("="):
                        # Use the Translator to adjust the formula for each subsequent row
                        for row_idx in range(start_row + 1, last_row + 1):
                            translator = Translator(formula, f"{column_letter}{start_row}")
                            adjusted_formula = translator.translate_formula(f"{column_letter}{row_idx}")
                            # Set the adjusted formula in the target row
                            sheet[f"{column_letter}{row_idx}"].value = adjusted_formula

                logger.info(f"The formulas were populated from sheet: {sheet_name}")
            else:
                logger.warning(f"Something went wrong when filling out the formulas !")
        
        
        # Save the changes to the file
        seed_wb.save(seed_copy)

        # Clean up by deleting the temporary files
        os.remove(file_path)
        
        # delete the file INPUT.xlsx if exists
        if os.path.exists(os.path.join(FOR_DOWNLOAD, "INPUT.xlsx")):
            # Delete the file
            os.remove(os.path.join(FOR_DOWNLOAD, "INPUT.xlsx"))

        return True
    
    def perform(self, 
                task_id: int,
                *args,
                **kwargs
                ) -> None:
        """
        This function executes the logic of the Import Task.
        """

        try:
            task = Task_CheckDbi.objects.get(pk=task_id)
        except ObjectDoesNotExist:
            print(
                f"Task with ID {task_id} was not found!"
            )
            raise
        
        task.start_date = timezone.now()
        task.status = TaskStatus.RUNNING
        task.save()

        logfile = pathlib.Path(task.logfile)
        result = False
        try:
            # create task's log directory
            logfile.parent.mkdir(parents=True, exist_ok=True)

            with Tee(logfile, "a"):
                result = self.execute(
                    task.id, # we send the task instance
                    *task.params.get("args", []),
                    **task.params.get("kwargs", {}),
                    )

        except Exception as exception:
            task.status = TaskStatus.FAILED
            task.save()

            traceback_info = "".join(
                traceback.TracebackException.from_exception(exception).format()
            )
            print(traceback_info)

            # try logging the exception
            try:
                with open(task.logfile, "a") as log:
                    log.write(traceback_info)
            except:
                pass
        else:
            task.status = TaskStatus.SUCCESS if result else TaskStatus.FAILED
            task.progress = 100
            task.save()
        finally:
            '''
            Final check of the ImportedSheet.
            If at least 1 sheet process is failed, the whole task is considered unsuccessful
            '''
            import_sheet = ImportedSheet.objects.filter(task_id__id=task.id)

            if len(import_sheet) > 0:
                imported_results = all(list(map(lambda x: x.status == 'SUCCESS', import_sheet)))
                task.status = TaskStatus.SUCCESS if imported_results else TaskStatus.FAILED

            task.progress = 100
            task.end_date = timezone.now()
            task.save()

    