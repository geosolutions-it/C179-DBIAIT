import os
import shutil
import pathlib
import json
import datetime
import pandas as pd
import gc

from django.utils import timezone
from django.db.models import ObjectDoesNotExist
from django.conf import settings

from openpyxl.formula.translate import Translator
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import numbers, PatternFill

from app.dbi_checks.models import Task_CheckDbi, ProcessState, TaskStatus, ProcessType
from app.dbi_checks.utils import YearHandler
from app.dbi_checks.tasks.checks_definitions.formulas_calc import CalcFormulas

import logging

logger = logging.getLogger(__name__)


class BaseCalc:

    def __init__(
        self, 
        orm_task,
        imported_file: str,
        seed: str,
        config: str,
        formulas_config: str,
        export_dir: pathlib.Path,
        file_year_required: bool = False,
        task_progress: int = 0,
        log_workbook = None
    ):
        """
        Initialization function of data export

        Parameters:
            export_dir: directory where the resulted xlsx files and log files should be stored
            orm_task: instance of the database Task reporting execution of this export task (default 100)
        """
        
        self.orm_task = orm_task
        self.imported_file = imported_file
        self.seed = seed
        self.config = config
        self.formulas_config = formulas_config
        self.export_dir = export_dir
        self.file_year_required = file_year_required
        self.task_progress = task_progress
        self.log_workbook = log_workbook

        self.logger = None

        # make sure target location exists
        self.export_dir.parent.mkdir(parents=True, exist_ok=True)

    def run(self):

        logger.info(f"exported dir path: {self.export_dir}")

        # It's crucial to use the read_only argument because it's quite faster
        up_file = load_workbook(self.imported_file, read_only=True)
        seed_basename = os.path.basename(self.seed)
        seed_copy = shutil.copy(self.seed, f"{self.export_dir}/{seed_basename}")
        seed_wb = load_workbook(seed_copy, data_only=False)

        logger.info(f"{self.file_year_required}")

        if self.file_year_required:
            # Create the INPUT.xlsx file which it is needed by the DBI_A formulas
            success = YearHandler(self.imported_file, self.export_dir).set_year_to_file()
            if success:
                logger.info("Year set successfully and INPUT.xlsx created.")
            else:
                logger.info("Failed to set the year or create INPUT.xlsx.")

        # Iterate over the sheets to copy data
        for source_sheet, config in self.config.items():
            
            start_date = timezone.now()
            
            target_sheet = config["target"]
            
            # Convert column letters to numbers
            start_col = column_index_from_string(config["start_col"])
            # Start row of the seed file
            target_start_row = config["start_row"]
            # Start row of the uploaded file
            source_start_row = config["source_start_row"]

            if source_sheet in up_file.sheetnames and target_sheet in seed_wb.sheetnames:

                source = up_file[source_sheet]
                target = seed_wb[target_sheet]

                # Copy data based on the specified column range
                for row_idx, row in enumerate(source.iter_rows(
                        min_row=source_start_row, 
                        max_row=source.max_row, 
                        min_col=start_col, 
                        max_col=source.max_column), 
                    start=target_start_row):
                    for cell in row:
                        if cell.value is not None:
                            # Calculate the target row position based on the row index
                            target_cell = target.cell(row=row_idx, column=cell.column, value=cell.value)

                            # If the value is a date, set date format
                            if isinstance(cell.value, datetime.date):
                                target_cell.number_format = 'MM/DD/YYYY'
                                target_cell.value = cell.value
                            # If the value is a number, set General format to avoid date misinterpretation
                            elif isinstance(cell.value, (int, float)):
                                target_cell.number_format = 'General'
                                target_cell.value = cell.value
                            else:
                                # For any other type (e.g strings), set the appropriate format
                                target_cell.number_format = '@'
                                target_cell.value = cell.value
  
                logger.info(f"Copied data from sheet: {source_sheet} to {target_sheet}")
                
                # Call the import_process_state function with the process type COPY and SUCCESS status
                end_date = timezone.now()
                self.import_process_state(self.orm_task.id, 
                                          ProcessType.COPY.value,
                                          os.path.basename(self.imported_file), 
                                          start_date, 
                                          end_date, 
                                          TaskStatus.SUCCESS,
                                          sheet=source_sheet
                                          )

            else:
                end_date = timezone.now()
                self.import_process_state(self.orm_task.id, 
                                          ProcessType.COPY.value, 
                                          os.path.basename(self.imported_file), 
                                          start_date, 
                                          end_date, 
                                          TaskStatus.FAILED,
                                          sheet=source_sheet
                                          )
                logger.warning(f"Sheet {source_sheet} or {target_sheet} not found!")
            
        self.orm_task.progress += self.task_progress
        self.orm_task.save()
    
        # Drag the formulas
        self.drag_formulas(seed_wb)

        # Write the year to the resulted file
        self.year_to_file(seed_wb)

        self.orm_task.progress += self.task_progress
        self.orm_task.save()

        start_date = timezone.now()
        logger.info(f"The file is ready to be saved")
        # save logic
        seed_wb.save(seed_copy)
        logger.info(f"Final workbook save completed.")
        end_date = timezone.now()

        # save the save process in the ProcessState model
        self.import_process_state(self.orm_task.id, 
                                  ProcessType.SAVE.value, 
                                  os.path.basename(self.imported_file), 
                                  start_date, 
                                  end_date, 
                                  TaskStatus.SUCCESS
                                  )
        # log file process
        self.log_file_manager(seed_wb, seed_basename)

        # Clean up by deleting the import file
        os.remove(self.imported_file)
        del seed_wb
        
        return True
    
    def drag_formulas(self, seed_wb):
        # Iterate through each sheet to drag the formulas
        for sheet_name, f_location in self.formulas_config.items():

            start_date = timezone.now()
            
            if sheet_name in seed_wb.sheetnames:
                sheet = seed_wb[sheet_name]

                # Get column indexes
                start_col_index = column_index_from_string(f_location["start_col"])
                end_col_index = column_index_from_string(f_location["end_col"])
                start_row = f_location["start_row"]
                # Re-definition of the last row because the copied file is processed
                # without saving yet. We don't want to re-load it for time reasons
                last_row = self.get_last_data_row(sheet)

                # Copy formulas from row 4 to the rest of the rows
                for col_idx in range(start_col_index, end_col_index + 1):
                    column_letter = get_column_letter(col_idx)
                    # Get the formula in row 4
                    formula = sheet[f"{column_letter}{start_row}"].value
                    if isinstance(formula, str) and formula.startswith("="):
                        # Use the Translator to adjust the formula for each subsequent row
                        for row_idx in range(start_row + 1, last_row + 1):
                            translator = Translator(formula, f"{column_letter}{start_row}")
                            adjusted_formula = translator.translate_formula(f"{column_letter}{row_idx}")
                            # Set the adjusted formula in the target row
                            target_cell = sheet[f"{column_letter}{row_idx}"]
                            target_cell.value = adjusted_formula
                            
                            # Explicit formatting
                            target_cell.number_format = numbers.FORMAT_GENERAL

                
                logger.info(f"The formulas were populated from sheet: {sheet_name}")
                # Call the import_process_state function with the process type CALCULATION and SUCCESS status
                end_date = timezone.now()
                self.import_process_state(self.orm_task.id, 
                                          ProcessType.CALCULATION.value,  
                                          os.path.basename(self.imported_file), 
                                          start_date, 
                                          end_date, 
                                          TaskStatus.SUCCESS,
                                          sheet=sheet_name
                                          )

            else:
                end_date = timezone.now()
                self.import_process_state(self.orm_task.id, 
                                          ProcessType.CALCULATION.value,
                                          os.path.basename(self.imported_file), 
                                          start_date, 
                                          end_date, 
                                          TaskStatus.FAILED,
                                          sheet=sheet_name
                                          )
                logger.warning(f"Something went wrong when filling out the formulas !")
    
    def import_process_state(self, task_id, process_type, file_name, start_date, end_date, status, sheet=""):
    
        try:
            task = Task_CheckDbi.objects.get(pk=task_id)

            ProcessState.objects.create(
                task=task,
                process_type = process_type,
                sheet_name=(sheet.lower() if sheet else ""),
                file_name=file_name,
                import_start_timestamp=start_date,
                import_end_timestamp=end_date,
                status=status
            )

            logger.info(f"The process {process_type} was completed successfully")
            return True

        except ObjectDoesNotExist:
            logger.error(f"Task with ID {task_id} was not found: {str(e)}")
            raise
        except Exception as e:
            logger.error(f"An error occurred while importing sheet: {str(e)}")
            raise
    
    def year_to_file(self, seed_wb):
        # Write the year to the resulted file
        defined_year = YearHandler(self.imported_file).get_year()
        dati_sheet = seed_wb["DATI"]
        dati_sheet['B8'] = defined_year
        logger.info(f"The year {defined_year} was copied to th DATI sheet")

    def year_to_file(self, seed_wb):
        # Write the year to the resulted file
        defined_year = YearHandler(self.imported_file).get_year()
        dati_sheet = seed_wb["DATI"]
        dati_sheet['B8'] = defined_year
        logger.info(f"The year {defined_year} was copied to th DATI sheet")
    
    def get_last_data_row(self, sheet):
        last_row = 0
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
           if any(cell.value is not None for cell in row):
                last_row = row[0].row
        return last_row
    
    def log_file_manager(self, seed_wb, seed_name):

        ## Configuration setup
        # prepare the logs workbook
        # Remove default sheet if it exists
        if "Sheet" in self.log_workbook.sheetnames and len(self.log_workbook.sheetnames) == 1:
            del self.log_workbook["Sheet"]
        
        sheet_name = "Logs"
        if sheet_name not in self.log_workbook.sheetnames:
            log_sheet = self.log_workbook.create_sheet(sheet_name)
            log_sheet.append(["File", 
                              "Foglio",
                              "Codice opera",
                              "Colonna check", 
                              "Tipo check", 
                              "Valore check errato col", 
                              "Valore errato col1", 
                              "Valore errato col2",
                              "Valore errato col3",
                              "Valore errato col4"

                              ])
        else:
            log_sheet = self.log_workbook[sheet_name]
        
        # Set the style in the log file
        self.set_logfile_style(log_sheet)
                
        # set the configs
        analysis_year = YearHandler(self.imported_file).get_year()

        # Get the seed_file name in order to retrieve it
        # from the log_mapping.json

        seed_filename = pathlib.Path(self.seed).stem
        seed_key = seed_filename.upper()
                
        with open(settings.LOG_MAPPING, "r", encoding='utf-8') as file:
            log_mapping = json.load(file)
                
        verif_checks_config = log_mapping.get(seed_key, {})

        formulas_config = self.load_formulas_conf(seed_key)
        
        ## Calculate the formulas of the checks for each sheet
        for sheet_name, f_location in formulas_config.items():

            pd_sheet = None
            start_date = timezone.now()
            
            if sheet_name in seed_wb.sheetnames:
                sheet = seed_wb[sheet_name]

                # Get column indexes
                start_col_index = column_index_from_string(f_location["start_col"])
                end_col_index = column_index_from_string(f_location["end_col"])
                start_row = f_location["start_row"]
                # Re-definition of the last row because the copied file is processed
                # without saving yet. We don't want to re-load it for time reasons
                end_row = self.get_end_row(f_location, sheet_name, sheet)
                
                calculator = self.get_calculator()
                # caclulate the formulas of the column checks
                sheet_with_calc_values = calculator(workbook=seed_wb, 
                                                sheet=seed_wb[sheet_name],
                                                start_row=start_row, 
                                                end_row = end_row,
                                                start_col = start_col_index,
                                                end_col = end_col_index,
                                                analysis_year=analysis_year,
                                                external_wb_path=self.export_dir,
                                                seed_name = seed_name
                                                ).main_calc()

                # Calculate the extra formulas of DB_prioritari
                # ........................................
                if "prioritari" in seed_name:
                    sheet_with_calc_values = self.calc_extra_prior_formulas(
                        calculator,
                        seed_name,
                        seed_wb,
                        sheet_name,
                        sheet_with_calc_values, 
                        analysis_year
                        )
                
                ## setup the config for each sheet
                sheet_checks = verif_checks_config.get(sheet_name, None)
                for check in sheet_checks:
        
                    # Get the verification check cell to check if it is OK or not
                    verif_check = check.get("verif_check", {})
                    if verif_check is None:
                        continue
                    verif_check_col = verif_check["col"]
                    verif_check_col_index = column_index_from_string(verif_check_col)
                    verif_check_row = verif_check["row"]

                    # caclulate the formula in the verification check
                    sheet_with_verif_values = calculator(workbook=seed_wb, 
                                                sheet=sheet_with_calc_values,
                                                start_row=verif_check_row, 
                                                end_row = verif_check_row,
                                                start_col = verif_check_col_index,
                                                end_col = verif_check_col_index,
                                                analysis_year=analysis_year,
                                                external_wb_path=self.export_dir,
                                                seed_name = seed_name,
                                                ).main_calc()
                    
                    # retrieve the calculated verif check value
                    verif_check_value = sheet_with_verif_values[f"{verif_check_col}{verif_check_row}"].value
                    if verif_check_value == "OK":
                        logger.info(f"The check of the cell {verif_check_col}{verif_check_row} is OK")
                            
                    else:
                        column_check = check.get("colonna_check", None)
                        check_name = check.get("check", None)                 
                        column_rel = check.get("colonna_rel", None)
                        
                        # convert column names for pandas
                        column_check_idx = self.parse_col_for_pd(column_check)
                        # column_rel_idx = self.parse_col_for_pd(column_rel)
                        
                        desc = check.get("descrizione", None)
                        criterion = check.get("valore", 0)
                        #logger.info(f"{check_name}, {column_rel}, {desc}")
                        
                        # Ensure column_rel is always a list (or empty if None)
                        if column_rel is None:
                            column_rel = []

                        if pd_sheet is None:
                            # Read all data from the worksheet
                            data = list(sheet.iter_rows(min_row=1, 
                                                        max_row=self.get_last_data_row(sheet_with_verif_values), 
                                                        values_only=True))  # Read all rows as tuples

                            # Convert to DataFrame
                            pd_sheet = pd.DataFrame(data)

                        # Get rows where column_check is NOT 0
                        start_idx = start_row - 1

                        #filtered_rows = pd_sheet.iloc[start_idx:][pd_sheet.iloc[start_idx:][column_check_idx] != criterion]
                        filtered_rows = pd_sheet.iloc[start_idx:][pd_sheet.iloc[start_idx:, column_check_idx] != criterion]
                        # Iterate through the filtered rows and retrieve the necessary information
                        for index, row in filtered_rows.iterrows():
                            incorrect_value = row[column_check_idx]  # Value of the cell in `colonna_check`
                            
                            # get the unique code (Codice opera)
                            unique_code = self.get_the_unique_code(sheet_name, row)
                            
                            # Handle the colonna_rel values:
                            # Initialize a dictionary to store values from each related column
                            related_values_dict = {}
                            for rel_col in column_rel:  # Iterate through all columns in column_rel list
                                # Ensure to get the value of each related column (if it exists)
                                if rel_col:
                                    related_value = row[self.parse_col_for_pd(rel_col)] if rel_col else None
                                    related_values_dict[rel_col] = related_value

                            # Replace placeholders in the description with the corresponding values from related_values_dict
                            # we re-define the updated_desc to be equal with the initial desc
                            updated_desc = desc
                            for key, value in related_values_dict.items():
                                placeholder = f"{{{key}}}"  # e.g., {AO}
                                if placeholder in updated_desc:
                                    updated_desc = updated_desc.replace(placeholder, str(value))

                            # Create a list for the related values, using the dictionary's get method to handle missing keys
                            related_values = [related_values_dict.get(key, None) for key in list(related_values_dict.keys())]

                            # Append the row to the log sheet
                            log_sheet.append([seed_key, sheet_name, unique_code, column_check, updated_desc, incorrect_value] + related_values)

                end_date = timezone.now()

                self.import_process_state(self.orm_task.id, 
                                        ProcessType.LOG.value,
                                        os.path.basename(self.imported_file), 
                                        start_date, 
                                        end_date, 
                                        TaskStatus.SUCCESS,
                                        sheet=sheet_name
                                        )
            
            # Remove DataFrame from memory and trigger garbage collection
            del pd_sheet
            gc.collect()
        
        self.task_progress = self.task_progress + 20

    def parse_col_for_pd(self, col):
        
        # Convert Excel-style column name (e.g., "BG") to 1-based index
        col_index = column_index_from_string(col)

        # Convert to 0-based index for pandas use
        col_index -= 1  # Convert to 0-based index
        return col_index
    
    def set_logfile_style(self, log_sheet):
        # Define a background color fill
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow color

        # Apply style to the first row
        for row in log_sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill

        # Set a default column width
        for col_cells in log_sheet.columns:
            col_letter = col_cells[0].column_letter
            log_sheet.column_dimensions[col_letter].width = 20

        # Set a much larger width for specific columns
        for col_letter in ["A", "E", "F", "G", "H", "I"]:
            log_sheet.column_dimensions[col_letter].width = 25

        # Set an even larger width for column D
        log_sheet.column_dimensions["D"].width = 45
    
    def get_the_unique_code(self, sheet_name, row):
        # retrieve the column B which includes the unique code of each record
        # we set all the suffixes of the sheet names that have the 
        # unique code in column A instead of B
        valid_suffixes = (
                          "_inpotab", 
                          "_inreti", 
                          "_tronchi", 
                          "_pompe", 
                          "_incaptaz", 
                          "_com_serv",
                          "_inadd", 
                          "_inserba", 
                          "_loc_serv", 
                          "_incoll", 
                          "_infog",
                         )
        
        if any(sheet_name.endswith(suffix) for suffix in valid_suffixes):
            # retrieve the column B which includes the unique code of each record
            unique_code_col = "A"
                             
        else:
            # retrieve the column B which includes the unique code of each record
            unique_code_col = "B"
        
        unique_code_idx = self.parse_col_for_pd(unique_code_col)
        unique_code = row[unique_code_idx]

        return unique_code
    
    def get_end_row(self, f_location, sheet_name, sheet):
        return self.get_last_data_row(sheet)
    
    def load_formulas_conf(self, seed_key):
        return self.formulas_config
    
    def get_calculator(self):
        return CalcFormulas
    
    def calc_extra_prior_formulas(self,
                                  calculator,
                                  seed_name,
                                  seed_wb,
                                  sheet_name,
                                  sheet_in_mem,
                                  analysis_year):
    
        '''
        This method calculates all the extra formulas that are
        existed in the DB prioritari file
        '''

        # Load the extra formulas config
        with open(settings.EXTRA_DB_PRIOR_FORMULAS, "r", encoding='utf-8') as file:
            extra_formulas = json.load(file)
        extra_formulas_config = extra_formulas.get(sheet_name, {})

        for conf in extra_formulas_config:
            # Get row / col references
            start_col_index = column_index_from_string(conf["start_col"])
            end_col_index = column_index_from_string(conf["end_col"])
            start_row = conf["start_row"]

            extra_calc_values = calculator(workbook=seed_wb, 
                                           sheet=sheet_in_mem,
                                           start_row=start_row, 
                                           end_row = start_row,
                                           start_col = start_col_index,
                                           end_col = end_col_index,
                                           analysis_year=analysis_year,
                                           external_wb_path=self.export_dir,
                                           seed_name=seed_name,
                                           ).main_calc()
        return extra_calc_values