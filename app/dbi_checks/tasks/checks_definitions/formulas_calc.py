import re
import math

import openpyxl.workbook

import formulas
import openpyxl
from openpyxl.utils.cell import column_index_from_string, get_column_letter



class CalcFormulas:
    """
    This class calculates different kinds of formulas
    using the Formulas module installed from here:
    https://github.com/geosolutions-it/formulas/tree/v1.2.8_fixed

    The formulas have to be parsed before being calculated.
    There are four kinds of groups:
    -- Simple formulas: They contain only built-in functions
        like COUNTIF and cells' references
    -- Formulas with ranges: They include also ranges like A:A
        which has to be calculated separately.
    -- Formulas with ranges from other sheet: They are similar
        with the formulas with ranges but in this case, the formulas
        take the data from other sheet.
    -- Formulas which retrieve the year value from the INPUT.xlsx file
        In our case we will just replace this file with the actual value
        because we have retrieve it in a previous step of the process.
    -- Formulas which retrieve data from other xlsx file. These formulas
        are included only in the DBI-A-1 file.
    """

    def __init__(self,
                 workbook: openpyxl.workbook.Workbook,
                 sheet: openpyxl.worksheet.worksheet.Worksheet,
                 start_row: int,
                 end_row: int,
                 start_col: int,
                 end_col:int,
                 analysis_year: int,
                 external_wb_path: str = None,
                 seed_name: str = None,
                 correct_values: list = None,
                 ):
        self.workbook = workbook
        self.sheet = sheet
        self.start_row = start_row
        self.end_row = end_row
        self.start_col = start_col
        self.end_col = end_col
        self.analysis_year = analysis_year
        self.external_wb_path = external_wb_path
        self.seed_name = seed_name
        self.correct_values = correct_values

    def main_calc(self):
        
        # A list which will include all the columns with incorrect values (which they are not OK)
        verif_checks_results = []
        
        # Iterate through the columns in the specified range
        for col_idx in range(self.start_col, self.end_col + 1):

            col_letter = get_column_letter(col_idx)

            # Get the correct value if it exists
            correct_value = self.correct_values.get(col_letter) if self.correct_values else None

            # Dictionary to set the variables of the formula
            variables = {}

            # Retrieve the formula from the first cell of the column
            formula_cell = self.sheet[f"{col_letter}{self.start_row}"]
                
            if formula_cell.data_type != 'f':  # Skip if not a formula
                continue
                
            formula = formula_cell.value

            if "prioritari" in self.seed_name:
                formula = self.replace_single_cell_counta(formula)
  
            # Parse and compile the formula outside the loop for better performance
            parser = formulas.Parser()
            ast = parser.ast(formula)[1]
            compiled = ast.compile()
            # Check if the formulas include the INPUT.xlsx requirement
            if re.search(r"\[\d+\]Input anno", formula, re.IGNORECASE):
                # Replace the entire external reference with analysis_year
                formula = self.replace_with_year(formula)
                ast = parser.ast(formula)[1]
                compiled = ast.compile()

            # Regular expression to extract the column letters from the formula (e.g., X, L) and sheet names if exist.
            # For instance, it will catch the cell A2 but also the Fiumi!A2.
            #columns_in_formula = re.findall(r'([A-Za-z_]+!)?([A-Z]{1,3})\d+', formula)
            # updated columns_in_formula in order to not catch the columns insite quotes: "A2"
            columns_in_formula = re.findall(r'(?<!["\'])([A-Za-z_]+!)?([A-Z]{1,3})\d+', formula)
            # Remove '!' from sheet names if present
            columns_in_formula = [(match[0].rstrip('!'), match[1]) for match in columns_in_formula]
            columns_in_formula = self.exclude_range_columns(columns_in_formula, formula)
                
            # Check if the formula includes ranges e.g A:A or sheet with ranges. It catch patterns like: Fiumi_inreti!A:A
            # which is the way that Openpyxl interprets internally the actual patterns e.g $Fiumi_inreti.A:A 
            # ranges_in_formula = re.findall(r'(?:(?P<sheet>[A-Za-z_][\w]*)!)?(?P<range>\$?[A-Z]{1,3}:\$?[A-Z]{1,3})', formula)
            # this new pattern catches ranges with and without rows like B1:B2232
            ranges_in_formula = re.findall(r'(?:(?P<sheet>[A-Za-z_][\w]*)!)?(?P<range>\$?[A-Z]{1,3}:\$?[A-Z]{1,3}|\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+)', formula)   
            # Regex to capture the ranges in one row e.g B4:BB4
            row_based_ranges = re.findall(r'(?P<col1>\$?[A-Z]{1,3})(?P<row>\d+):(?P<col2>\$?[A-Z]{1,3})(?P=row)', formula)

            # Regex to capture the absolute rows like B$1 (They are used in the DB_prioritari file)
            abs_rows = re.findall(r'\b[A-Z]+\$\d+\b', formula)

            # Regex to capture the fake relative cells. The main verification formula of DB_prioritari includes
            # cell reference in the row 1 or 2 e.g A1 where it is absolute. So, for prefenting our calculator
            # to iterate through all the rows, we have to catch all the cells in rows 1 and 2 and exclude them
            # from columns_in_formula
            fake_rel_cells = re.findall(r'\b([A-Z]{1,3})([12])\b(?!:)', formula) # it outputs a format like: [('AC', '2'), ('AE', '1')]
          
            if ranges_in_formula:
                for match in ranges_in_formula:
                    sheet_name, col_ranges = match
                    # remove the $ from the col_ranges if exist
                    col_ranges = col_ranges.replace('$', '')
                    # Skip row-based ranges (B4:BB4, A10:C10)**

                    # The check below removes the row-based ranges if there are in the ranges_in_formula
                    if re.match(r'^(?P<col1>[A-Z]{1,3})(?P<row>\d+):(?P<col2>[A-Z]{1,3})(?P=row)$', col_ranges):
                        continue  # Skip this range, go to next match
                    
                    if sheet_name:
                        # check if the sheet is from another file, In this case the interpreted formula will be like: [2]FIUMI!B:B
                        second_file_check = re.search(r'\[(\d+)\]', formula)
                        if second_file_check:
                            linked_number = second_file_check.group(1)
                            external_wb = self.external_link_parser(linked_number)
                           
                            col_ranges = re.sub(r'(\d+)(?=\$?\d*$)', '', col_ranges)  # Remove row number after the first column
                            formatted_variable = f"[{linked_number}]{sheet_name.upper()}!{col_ranges}"
                            variables[formatted_variable] = self.calculate_range(formula, col_ranges, sheet_name, external_wb)
                        else:
                            formatted_variable = f"{sheet_name.upper()}!{col_ranges}"  # e.g FIUMI_INRETI!A:A
                            variables[formatted_variable] = self.calculate_range(formula, col_ranges, sheet_name)
                    else:
                        variables[col_ranges] = self.calculate_range(formula, col_ranges)
                
            if row_based_ranges:
                columns_in_formula = self.exclude_cols_from_row_ranges(row_based_ranges, columns_in_formula)
                for i in row_based_ranges:
                    # the structure of row_based_ranges is: [(col1, row, col2)]
                    variables[f"{i[0]}{i[1]}:{i[2]}{i[1]}"] = self.calculate_row_based_range(i) # e.g variables[B4:BB4]
            if abs_rows:
                abs_rows = self.exclude_abs_range_columns(abs_rows, ranges_in_formula)
                for i in abs_rows:
                    # the format of the abs_rows is something like "['J$1']"
                    i = i.replace("$", "")
                    variables[i] = self.sheet[f"{i}"].value
            
            if "prioritari" in self.seed_name:
                if fake_rel_cells:
                    # exclude fake_rel_cells from columns_in_formula
                    fake_cols_set = {col for col, row in fake_rel_cells}  # {'AC', 'AE'}
                    columns_in_formula = [match for match in columns_in_formula if match[1] not in fake_cols_set]
                    for col, row in fake_rel_cells:
                        variables[f"{col}{row}"] = self.sheet[f"{col}{row}"].value

            # Iterate through each row for this column
            for row in self.sheet.iter_rows(min_row=self.start_row, max_row=self.end_row,
                                            min_col=col_idx, max_col=col_idx):
                cell = row[0]
                    
                # Retrieve the required values from the relevant cells
                for sheet_name, col in columns_in_formula:
                    ref_cell = f"{col}{cell.row}"
                    if not sheet_name:
                        value = self.sheet[ref_cell].value
                        # We set the text values in uppercase:
                        value = self.cell_value_parser(value)
                        # We set the key of the variables using the first row with data (4)
                        # because in that way have been compliled by Formulas. The result
                        # of course is updated with the new rows.
                        variables[f"{col}{self.start_row}"] = self.sanitize_value(value, formula)  # Default to 0 if empty
                    else:
                        value = self.workbook[sheet_name][ref_cell].value
                        # We set the text values in uppercase:
                        value = self.cell_value_parser(value)
                        
                        variables[f"{sheet_name.upper()}!{col}{self.start_row}"] = self.sanitize_value(value, formula)  # Default to 0 if empty

                # Evaluate the formula with the given variables
                try:
                    calculated_result = compiled(**variables)
                except Exception as e:
                    calculated_result = f"Error: {e}"

                result = calculated_result.item() if hasattr(calculated_result, "item") else calculated_result
                
                # Check if the result is the correct value in case of the column checks
                if correct_value is not None:
                    if result != correct_value:
                        # store this as 1 which means that this column is not OK
                        verif_checks_results.append(col_letter)
                
                # convert the float to int if the result is float
                #if isinstance(result, float):
                #    result = int(round(result))

                # print(f"Sheet: {self.sheet}, Row {cell.row} ({col_letter}{cell.row}): {result}")
                # Store the result in the target cell
                cell.value = result

        # return the caclulated_result as single value and not as a numpy array e.g Array("OK", dtype=object)
        return (self.sheet, verif_checks_results)

    def calculate_range(self, formula: str,  col_range: str, sheet_name: str = None, external_wb: openpyxl.workbook.Workbook = None) -> list:
        """
        Calculate the values in a given Excel range either from the initial sheet
        or from another one
        """
        
        # define the first row of the range:
        start_row = self.get_the_first_row_from_a_range(col_range)

        vlookup_pattern = r'\bVLOOKUP\('

        if external_wb:
            current_sheet = external_wb[sheet_name]
        else:
            # Use the specified sheet or default to the current one
            current_sheet = self.sheet if not sheet_name else self.workbook[sheet_name]
        
        # Remove dollar signs for absolute references
        col_range = col_range.replace('$', '')

        # TODO See if we indeed need this function
        col_range = self.clean_range(col_range)
        
        # Split start and end of the range
        start_col, end_col = col_range.split(':')
        
        # Convert column letters to column numbers
        start_num = column_index_from_string(start_col)
        end_num = column_index_from_string(end_col)
        
        # Create an empty list to hold the values
        values = []
        
        # Iterate through the range of columns (same row range for each column)
        for col_num in range(start_num, end_num + 1):
            col_letter = get_column_letter(col_num)
            # Get values in the specified column
            column_values = [current_sheet[f'{col_letter}{row}'].value for row in range(start_row, self.get_last_data_row(current_sheet) + 1)]
            
            values.append(column_values)

        # Check if VLOOKUP exists in the formula first
        if re.search(vlookup_pattern, formula, re.IGNORECASE):
            # We need to build a 2D list in order to be interpretable by VLOOKUP
            # e.g [[search_col_value1, return_col_value1], 
            #      [search_col_value2, return_col_value2]
            #      ...]
            values = [list(tup) for tup in zip(*values)]
        
        return values
    
    def exclude_range_columns(self, columns_in_formula, formula):
        """
        This method exclude from the columns_in_formula the columns
        that have been added but they are part of a range and not
        actual columns like BG6:BG10
        """
        cleaned_columns = []
        for match in columns_in_formula:
            sheet_prefix, column = match  # match is a tuple (sheet name, column)
            
            # Check if the column is part of a range by looking for "column+number:" pattern
            if not re.search(rf'{column}\d+:\b', formula):  
                cleaned_columns.append((sheet_prefix, column))
        return cleaned_columns
    
    def exclude_abs_range_columns(self, abs_rows, ranges_in_formula):
        """
        This method excludes from the abs_rows the columns that are part of a range
        and not actual absolute columns (like B$2 but not B$1:B$1048576).
        """
        cleaned_columns = []
        
        # Flatten the ranges_in_formula into a list of columns (excluding the sheet names)
        columns_in_ranges = set()
        
        for sheet, range_expr in ranges_in_formula:
            # Ensure the range expression is valid and not empty
            if range_expr and ":" in range_expr:
                start_col, end_col = range_expr.split(':')
                
                # Ensure both parts are valid column references
                if '$' in start_col and '$' in end_col:
                    start_column = start_col.split('$')[1]  # Get the column part, e.g., "B" from "$B$1"
                    end_column = end_col.split('$')[1]      # Get the column part, e.g., "B" from "$B$1048576"
                    
                    # Add both start and end columns to the set
                    columns_in_ranges.add(start_column)
                    columns_in_ranges.add(end_column)
                else:
                    # If the range is invalid or does not have valid columns, skip it
                    continue

        # For each absolute row found in the formula
        for match in abs_rows:
            column = match.split('$')[0]  # Extract the column part of the match (e.g., "B" from B$1)
            
            # If the column is not part of any range, add it to the cleaned list
            if column not in columns_in_ranges:
                cleaned_columns.append(match)
        
        return cleaned_columns

    def clean_range(self, col_range: str) -> str:
        """
        Remove row numbers from a given range and return only the column part.
        Example: "$B$1:$B$10" -> "$B:$B"
        """

        if ':' in col_range:
            start_cell, end_cell = col_range.split(':')
            
            # Use regex to remove row numbers, keeping only column letters
            start_col = re.sub(r'\d+', '', start_cell)
            end_col = re.sub(r'\d+', '', end_cell)
            
            # If the start and end columns are the same, return just one column reference
            if start_col == end_col:
                return f"{start_col}:{start_col}"
            else:
                return f"{start_col}:{end_col}"
        
        # If it's a single cell or range without row numbers, just return it as-is
        return re.sub(r'\d+', '', col_range)

    def calculate_row_based_range(self, row_range):
        """
        Given a row-based range like ('B', '4', 'BB'),
        this function returns a list of all (column, row) tuples.
        """
        start_col, row, end_col = row_range  # Unpack the values
        row = int(row)  # Ensure row is an integer

		# Convert column letters to numerical indices
        start_index = column_index_from_string(start_col)
        end_index = column_index_from_string(end_col)

		# Generate all column names between start and end
        result = [
		    self.sheet[f"{get_column_letter(col)}{row}"].value
		    for col in range(start_index, end_index + 1)
		]

        return result
    
    def exclude_cols_from_row_ranges(self, row_based_ranges, columns_in_formula):
        excluded_columns = set()

        # Process the row-based ranges
        for start_col, row, end_col in row_based_ranges:
            # Add only the start and end columns to excluded_columns
            excluded_columns.add(start_col)
            excluded_columns.add(end_col)

        # Extract just the column names from columns_in_formula (which are the second element of the tuples)
        # Iterate over columns_in_formula and check if the column (second element) is in excluded_columns
        valid_columns = [
            (sheet_name, col) for sheet_name, col in columns_in_formula  # Keep the tuple structure intact
            if col not in excluded_columns  # Exclude columns in the row-based range
        ]

        return valid_columns

    def replace_with_year(self, formula):
        return re.sub(
                   r"'?\[\d+\]Input anno'!\$[A-Z]+\$[0-9]+",
                   f"{self.analysis_year}",
                   formula
                   )
    
    def get_last_data_row(self, sheet):
        last_row = 0
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
           if any(cell.value is not None for cell in row):
                last_row = row[0].row
        return last_row
    
    def external_link_parser(self, linked_number):
        
        # insert here the possible external files
        target_file = "DBI_A.xlsx"

        # Step 1: Extract all external links and map them to indices
        external_links_map = {}  # Stores {index_number: file_path}
        if self.workbook._external_links:
            for idx, link in enumerate(self.workbook._external_links, start=1):  # Excel indexes start at 1
                external_links_map[idx] = link.file_link.Target  # Map [1], [2], etc. to their file path

        # Step 2: Check if a linked number e.g [2] exists and matches the target file
        if int(linked_number) in external_links_map:
            if target_file in external_links_map[int(linked_number)]:
                #print(f"{linked_number} correctly links to {target_file}")
                # load the external workbook
                external_wb = openpyxl.load_workbook(f'{self.external_wb_path}/DBI_A.xlsx', data_only=True)

                return external_wb
        else:
            #print("[2] does not exist in the external links.")
            return False

    def cell_value_parser(self, value):
         if isinstance(value, str):
             value = value.upper()
         return value

    def get_the_first_row_from_a_range(self, range):
        '''
        This method returns the first row from a range.
        if we have a range like B5:B1000, will take return
        the value 5 as start row. Otherwise will return the 
        the default self.start_row
        '''
        match = re.match(r'([A-Z]{1,3})(\d+):[A-Z]{1,3}(\d+)', range)
        if match:
            col1, row1, row2 = match.groups()
            if row1 != row2:  # Check if row numbers are different
                return int(row1)
        else:
            return self.start_row
        
    def replace_single_cell_counta(self, formula: str) -> str:
        """
        Replaces occurrences of COUNTA(cell) with LEN(cell) when COUNTA is applied to a single cell.
        e.g: '=IF(AND(AR6<3,COUNTA(B6)>0),1,0)' → '=IF(AND(AR6<3,LEN(B6)>0),1,0)'
        """
        pattern = r"COUNTA\(\s*([A-Z]+\d+)\s*\)"

        # Replace with LEN(cell)
        return re.sub(pattern, r"LEN(\1)", formula)
    
    def sanitize_value(self, value, formula):
        '''
        This method handles the empty cells like the Excel
        '''
        if value is None or (isinstance(value, float) and math.isnan(value)) or value == "":
            return 0

        return value