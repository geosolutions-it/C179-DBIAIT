import math
import logging
import pathlib
import openpyxl

from datetime import datetime
from openpyxl.utils import cell

from django.conf import settings
from django.db import connection, connections, ProgrammingError

from .domains_parser import Domains
from .config_scraper import ExportConfig

from app.scheduler.models import Task
from app.scheduler.utils import dictfetchall, translate_schema_to_db_alias


class ExportXls:

    SEED_FILE_ID_ROW = 3

    def __init__(
        self, export_dir: pathlib.Path, orm_task: Task, max_progress: int = 100
    ):
        """
        Initialization function of data export to *.xlsx file

        Parameters:
            export_dir: directory where export and log files should be stored
            orm_task: instance of the database Task reporting execution of this export task (default 100)
            max_progress: max value of Task's progress, which should be set after a successful export
        """
        self.export_dir = export_dir
        self.orm_task = orm_task
        self.max_progress = max_progress

        # make sure target location exists
        self.export_dir.parent.mkdir(parents=True, exist_ok=True)

    def configure_file_logger(self, logfile_path: pathlib.Path):
        """
        Method configuring logger for logging user dedicated errors of the export into a specified location.

        Parameters:
            logfile_path: path of the log file
        """
        logger = logging.getLogger(__name__)
        hdlr = logging.FileHandler(logfile_path.absolute())
        formatter = logging.Formatter("%(asctime)s %(levelname)s %(message)s")
        hdlr.setFormatter(formatter)
        logger.addHandler(hdlr)
        logger.setLevel(logging.INFO)

        return logger

    def run(self):
        """
        Method executing export of the data into *.xlsx file,
        """
        starting_progress = self.orm_task.progress
        today = datetime.today()

        # create log file
        logfile_path = pathlib.Path(
            self.export_dir.absolute(), f"logfile_{today.strftime('%Y%m%d')}.log"
        )
        logger = self.configure_file_logger(logfile_path)

        # get target xls file location
        target_xls_file = pathlib.Path(
            self.export_dir.absolute(),
            f"{today.strftime('%Y%m%d')} - NETSIC_{today.strftime('%Y')}.xlsx",
        )

        # parse export configuration
        config = ExportConfig()

        # calculate total number of steps
        total_sheet_number = len(config)
        step = 1

        # fetch all_domains info
        all_domains = Domains()

        # load seed *.xlsx file
        excel_wb = openpyxl.load_workbook(settings.EXPORT_XLS_SEED_FILE.substitute())

        for sheet in ExportConfig():

            # execute pre_process for the sheet
            pre_process = sheet.get("pre_process", None)
            if pre_process is not None:
                with connection.cursor() as cursor:
                    try:
                        cursor.callproc(f"{self.orm_task.schema}.{pre_process}")
                    except Exception as e:
                        logger.error(
                            f"Procedure '{pre_process}' called by sheet '{sheet['sheet']}' FAILED with:\n"
                            f"{type(e).__name__}: {e}.\n"
                            f"Skipping '{sheet['sheet']}' sheet generation."
                        )
                        continue
                    else:
                        result = cursor.fetchone()
                        logger.debug(
                            f"Procedure '{pre_process}' called by sheet '{sheet['sheet']}' executed: {result}."
                        )

            # set current sheet as active in the xls workbook
            try:
                sheet_index = excel_wb.sheetnames.index(sheet["sheet"])
            except ValueError:
                logger.error(
                    f"Seed file does not contain '{sheet['sheet']}' sheet. Skipping..."
                )
                continue

            excel_wb.active = sheet_index
            excel_ws = excel_wb.active

            # create sheet's column ID <-> column ID mapping (iterate over ID row: 3)
            coord_id_mapping = {}

            for column_index in range(1, len(excel_ws[self.SEED_FILE_ID_ROW]) + 1):
                column_letter = cell.get_column_letter(column_index)
                coord_id_mapping.update(
                    {
                        str(
                            excel_ws[f"{column_letter}{self.SEED_FILE_ID_ROW}"].value
                        ).strip(): column_letter
                    }
                )

            # get the index of the first empty excel row
            first_empty_row = max(len(col) for col in excel_ws.iter_cols()) + 1

            with connections[
                translate_schema_to_db_alias(self.orm_task.schema)
            ].cursor() as cursor:
                sql_sources = sheet["sql_sources"]

                if not sql_sources:
                    logger.warning(
                        f"Sources for '{sheet['sheet']}' is empty. Skipping..."
                    )
                    continue

                raw_data = []
                for source in sql_sources:
                    try:
                        cursor.execute(source)
                    except ProgrammingError as e:
                        logger.error(
                            f"Fetching source for sheet '{sheet['sheet']}' failed with:\n"
                            f"{type(e).__name__}: {e}.\n"
                            f"Source: '{source}'."
                        )
                        continue

                    raw_data.extend(dictfetchall(cursor))

            for raw_data_row in raw_data:
                # prepare data to be inserted into excel
                sheet_row = {}

                for column in sheet["columns"]:
                    try:
                        transformed_value = column["transformer"].apply(
                            row=raw_data_row, domains=all_domains
                        )
                    except Exception as e:
                        logger.error(
                            f"Error occurred during transformation of column with "
                            f"ID '{column['id']}' in row '{first_empty_row}' in sheet '{sheet['sheet']}':\n"
                            f"{type(e).__name__}: {e}.\n"
                        )
                        transformed_value = None

                    sheet_row.update({column["id"]: transformed_value})

                    for validator in column.get("validators", []):
                        if not validator["validator"].validate(transformed_value):
                            message = validator.get("warning", "")
                            column_letter = coord_id_mapping.get(
                                str(column["id"]), None
                            )

                            if message:
                                message = (
                                    message.replace("{SHEET}", sheet["sheet"])
                                    .replace("{ROW}", first_empty_row)
                                    .replace("{FIELD}", column_letter)
                                )
                                logger.error(message)
                            else:
                                logger.error(
                                    f"Validation failed for cell '{column_letter}{first_empty_row}' "
                                    f"in the '{sheet['sheet']}' sheet."
                                )

                # insert sheet_row into excel
                for column_id, value in sheet_row.items():
                    column_letter = coord_id_mapping.get(str(column_id))
                    if not column_letter:
                        logger.error(
                            f"No column with ID '{column_id}' found in '{sheet['sheet']}' sheet."
                        )
                        continue

                    try:
                        excel_ws[f"{column_letter}{first_empty_row}"] = value
                    except Exception as e:
                        logger.error(
                            f"Error occurred during inserting value to the column with "
                            f"ID '{column_id}' in row '{first_empty_row}' in sheet '{sheet['sheet']}':\n"
                            f"{type(e).__name__}: {e}.\n"
                        )

                first_empty_row += 1

            # update task status
            step += 1
            self.orm_task.progress = math.floor(
                step * (self.max_progress - starting_progress) / total_sheet_number
            )
            self.orm_task.save()

        # save updated *.xlsx seed file in the target location
        excel_wb.save(target_xls_file)

        self.orm_task.progress = self.max_progress
        self.orm_task.save()
