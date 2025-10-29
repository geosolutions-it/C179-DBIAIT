import pathlib
import re

import openpyxl

from datetime import datetime
from openpyxl.utils import cell

from django.conf import settings
from django.db import connections, ProgrammingError

from app.scheduler.utils import dictfetchall, translate_schema_to_db_alias, translate_schema_to_enum
from app.scheduler.tasks.export_definitions.domains_parser import Domains
from app.scheduler.tasks.export_definitions.municipalities_parser import Municipalities
from app.scheduler.tasks.export_definitions.config_scraper import XlsExportConfig

from .export_base import ExportBase
from .post_validation import PostValidation


class ExportXls(ExportBase):

    SEED_FILE_ID_ROW = 3

    def run(self):
        """
        Method executing export of the data into *.xlsx file,
        """
        # update starting progress
        self.starting_progress = self.orm_task.progress
        today = datetime.today()
        self.configure_file_logger()

        # get target xls file location
        target_xls_file = pathlib.Path(
            self.export_dir.absolute(),
            f"{today.strftime('%Y%m%d')} - NETSIC_{today.strftime('%Y')}.xlsx",
        )

        # parse export configuration
        config = XlsExportConfig(self.ref_year)

        # calculate total number of steps
        print(f"Exporting XLS for {len(config)} config files")
        total_sheet_number = len(config)
        step = 1

        # fetch all_domains info
        all_domains = Domains(schema=translate_schema_to_enum(self.orm_task.schema), year=self.ref_year)
        municipalities = Municipalities(schema=translate_schema_to_enum(self.orm_task.schema), year=self.ref_year)

        # load seed *.xlsx file
        seed_path = settings.EXPORT_XLS_SEED_FILE.substitute()
        if self.ref_year is not None:
            seed_path = settings.EXPORT_XLS_SEED_FILE.substitute({"year": self.ref_year})
        excel_wb = openpyxl.load_workbook(seed_path)

        for sheet in config:

            # execute pre_process for the sheet
            pre_process = sheet.get("pre_process", None)
            try:
                self.execute_pre_process(pre_process)
            except Exception as e:
                self.logger.error(
                    f"Procedure '{pre_process}' called by sheet '{sheet['sheet']}' FAILED with: "
                    f"{type(e).__name__}: {e}. "
                    f"Skipping '{sheet['sheet']}' sheet generation."
                )
                continue

            # set current sheet as active in the xls workbook
            try:
                sheet_index = excel_wb.sheetnames.index(sheet["sheet"])
            except ValueError:
                self.logger.error(
                    f"Seed file does not contain '{sheet['sheet']}' sheet. Skipping..."
                )
                continue

            excel_wb.active = sheet_index
            excel_ws = excel_wb.active

            # create sheet's column ID <-> column ID mapping (iterate over ID row: 3)
            coord_id_mapping = {}

            for column_index in range(1, len(excel_ws[self.SEED_FILE_ID_ROW]) + 1):
                column_letter = cell.get_column_letter(column_index)
                coord_id_mapping.update(
                    {
                        str(
                            excel_ws[f"{column_letter}{self.SEED_FILE_ID_ROW}"].value
                        ).strip(): column_letter
                    }
                )

            # get the index of the first empty excel row
            first_empty_row = 4

            with connections[
                translate_schema_to_db_alias(self.orm_task.schema)
            ].cursor() as cursor:
                sql_sources = sheet["sql_sources"]

                if not sql_sources:
                    self.logger.warning(
                        f"Sources for '{sheet['sheet']}' is empty. Skipping..."
                    )
                    continue

                raw_data = []
                for source in sql_sources:
                    try:
                        cursor.execute(source)
                    except ProgrammingError as e:
                        self.logger.error(
                            f"Fetching source for sheet '{sheet['sheet']}' failed with: "
                            f"{type(e).__name__}: {e}. "
                            f"Source: '{source}'."
                        )
                        continue

                    raw_data.extend(dictfetchall(cursor))

            for raw_data_row in raw_data:
                # prepare data to be inserted into excel
                sheet_row = {}

                for column in sheet["columns"]:
                    message = "Foglio: {SHEET}, Riga:{ROW}, Codice_ato: {CODICE_ATO}, Campo: {FIELD}: Error: {E}" if not column['warning'] else column['warning']
                    try:
                        transformed_value = column["transformer"].apply(
                            row=raw_data_row, domains=all_domains, municipalities=municipalities, already_transformed_data=sheet_row
                        )
                    except KeyError as e:
                        transformed_value = None
                    except TypeError as e:
                        transformed_value = None
                    except Exception as e:
                        warning_log = message or "Foglio: {SHEET}, Riga:{ROW}, Codice_ato: {CODICE_ATO}, Campo: {FIELD}: Transformation error"

                        warning_to_log = (
                            warning_log.replace("{SHEET}", sheet["sheet"])
                            .replace("{ROW}", str(first_empty_row))
                            .replace("{FIELD}", column.get("alias", column['id']))
                            .replace("{CODICE_ATO}", raw_data_row.get("codice_ato", "") or "")
                            .replace("{E}", e.args[0].strip('\n'))
                            .replace("{REF_YEAR}", str(self.ref_year or datetime.today().year))
                        )

                        if '{custom:' in warning_log:
                            re_pattern = re.compile('.*{custom:(.*)}\|Campo')
                            custom_field_name = re.match(re_pattern, message).group(1)
                            custom_field_value = raw_data_row.get(custom_field_name, "")
                            warning_to_log = warning_to_log.replace("{custom:" + custom_field_name + "}|",
                                                                    f"{custom_field_name.upper()}: {custom_field_value},")

                        self.logger.error(warning_to_log)
                        #  self.logger.error(
                        #      f"Trasformazione: Error occurred during transformation of column with "
                        #      f"ID '{column['id']}' in row '{first_empty_row}' in sheet '{sheet['sheet']}':\n"
                        #      f"{type(e).__name__}: {e}.\n"
                        #  )
                        transformed_value = None

                    sheet_row.update({column["id"]: transformed_value})

                for column in sheet["columns"]:
                    for validator in column.get("validators", []):
                        try:
                            if not validator["validator"].validate(sheet_row, self.ref_year):
                                message = validator.get("warning", "")
                                # column_letter = coord_id_mapping.get(
                                #     str(column["id"]), None
                                # )
                                # {custom:codice_ato} ->

                                warning_log = message or "Foglio: {SHEET}, Riga:{ROW}, Codice ato: {CODICE_ATO}, Codice_ato: {CODICE_ATO}, Campo: {FIELD}: Validation error"

                                warning_to_log = warning_log.replace("{SHEET}", sheet["sheet"])\
                                    .replace("{ROW}", str(first_empty_row))\
                                    .replace("{FIELD}", column.get("alias", column['id']))\
                                    .replace("{CODICE_ATO}", raw_data_row.get("codice_ato", "") or "")\
                                    .replace("{REF_YEAR}", str(self.ref_year or datetime.today().year))

                                if '{custom:' in warning_log:
                                    re_pattern = re.compile('.*{custom:(.*)}\|Campo')
                                    custom_field_name = re.match(re_pattern, message).group(1)
                                    custom_field_value = raw_data_row.get(custom_field_name, "")
                                    warning_to_log = warning_to_log.replace("{custom:" + custom_field_name + "}|", f"{custom_field_name.upper()}: {custom_field_value}|")

                                self.logger.warning(warning_to_log)

                        except Exception as e:
                            self.logger.error(
                                f"Validation: Error occurred during validation of column with "
                                f"ID '{column['id']}' in row '{first_empty_row}' in sheet '{sheet['sheet']}': "
                                f"{type(e).__name__}: {e.args[0]}.".strip("\n")
                            )
                # insert sheet_row into excel
                for column_id, value in sheet_row.items():
                    column_letter = coord_id_mapping.get(str(column_id))
                    if not column_letter:
                        self.logger.error(
                            f"No column with ID '{column_id}' found in '{sheet['sheet']}' sheet."
                        )
                        continue

                    try:
                        excel_ws[f"{column_letter}{first_empty_row}"] = value
                    except Exception as e:
                        self.logger.error(
                            f"Excel Config: Error occurred during inserting value to the column with "
                            f"ID '{column_id}' in row '{first_empty_row}' in sheet '{sheet['sheet']}': "
                            f"{type(e).__name__}: {e.args[0]}.".strip("\n")
                        )

                first_empty_row += 1

            # update task status
            step += 1
            self.update_progress(step, total_sheet_number)

        #  update the information in the sheet "DATI" before save it
        # Validation between sheets

        excel_wb["DATI"]["B5"] = today.date()
        excel_wb["DATI"]["B8"] = self.ref_year or datetime.utcnow().year
        excel_wb["DATI"]["B10"] = today.date()
        # save updated *.xlsx seed file in the target location
        excel_wb.save(target_xls_file)

        import json
        configs = []
        year = self.ref_year if self.ref_year else 'current'
        with open(f'{settings.EXPORT_CONF_DIR}/{year}/sheet_validation.json', encoding='utf-8') as file:
            configs = json.load(file)
        for val in configs:
            getattr(PostValidation, val['function'])(val, excel_wb, self.logger)

        self.set_max_progress()

