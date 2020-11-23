import json

from django.core.management.base import BaseCommand
from collections import OrderedDict
import os
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Given a specific xlsx into json"
    configuration_dict = {}

    def add_arguments(self, parser):
        parser.add_argument("abs_file_path", nargs="+", type=str)

    def handle(self, *args, **options):
        for file_path in options["abs_file_path"]:
            self.stdout.write(f"Start importing file: {file_path}")
            output = []
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb["SCHEMA"]
                output = self.gather_info_from_file(output, wb, ws)
                aggregation = self.aggregate_by_sheet(output)
                print(aggregation)
                self.stdout.write("File successfully loaded")

                self.generate_output(aggregation)

            else:
                self.stdout.write(
                    f"File does not exists, please check the file_path: {file_path}"
                )

    @staticmethod
    def aggregate_by_sheet(output):
        result = OrderedDict()
        for field, sheet, cid in output:
            if sheet in result:
                if "none.none" in field or " .none" in field or " . " in field:
                    pass
                else:
                    result[sheet].append((cid, field))
            else:
                if "none.none" in field or " .none" in field or " . " in field:
                    result[sheet] = []
                else:
                    result[sheet] = [(cid, field)]
        return [{"sheet": k, "fields": v} for k, v in result.items()]

    @staticmethod
    def gather_info_from_file(output, wb, ws):
        for index, item in enumerate(ws["A3":"A750"], start=3):
            sheet_name = item[0].value
            field_name = ws[f"E{index}"].value
            table_name = ws[f"F{index}"].value
            column_id = ws[f"D{index}"].value
            output.append(
                (
                    f"{table_name}.{field_name}".lower(),
                    sheet_name.lower().capitalize(),
                    str(int(column_id)),
                )
            )
        wb.close()
        return output

    def generate_output(self, aggregation: list):
        output_path = os.path.dirname(os.path.abspath(__file__))
        for configuration in aggregation:
            filename = f'{output_path}/conversion_output/{configuration["sheet"]}'
            with open(f"{filename}.json", "w+") as file:
                output = self.__output_structure()
                output["sheet"] = configuration["sheet"]
                output["sources"][0]["fields"] = [
                    {"name": x[1], "alias": x[1].split(".")[1]}
                    for x in configuration["fields"]
                ]
                output["columns"] = [
                    {
                        "id": x[0],
                        "transformation": {
                            "func": "",
                            "params": {"field": x[1].split(".")[1]},
                        },
                    }
                    for x in configuration["fields"]
                ]
                file.write(json.dumps([output], indent=4))

    @staticmethod
    def __output_structure():
        return {
            "sheet": "",
            "skip": True,
            "sources": [
                {
                    "fields": [{}],
                    "table": {"name": ""},
                    "join": [
                        {"type": "", "table": {"name": ""}, "on": ["", ""], "cond": ""}
                    ],
                    "filter": "",
                },
                {
                    "table": {"name": ""},
                    "join": [
                        {"type": "", "table": {"name": ""}, "on": ["", ""], "cond": ""}
                    ],
                    "filter": "",
                },
            ],
            "columns": [{}],
        }
