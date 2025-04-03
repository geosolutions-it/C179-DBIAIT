import re
import math

import pandas as pd

import formulas
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils import column_index_from_string
import openpyxl.workbook

from django.db.models import ObjectDoesNotExist
from django.utils import timezone

from app.dbi_checks.models import TaskStatus, ProcessType
from app.shape_checks.models import Task_CheckShape, ShapeCheckProcessState
from app.dbi_checks.tasks.checks_definitions.formulas_calc import CalcFormulas
import logging

logger = logging.getLogger(__name__)


class ImportStateHelper:
    """
    Imports the column state for a given task.

    :param task_id: ID of the task to update.
    :param process_type: Type of the process.
    :param start_date: Start timestamp of import.
    :param end_date: End timestamp of import.
    :param status: Status of the import.
    :param sheet: Optional sheet name.
    :return: True if successful, raises an exception otherwise.
    """
    @staticmethod
    def import_column_state(task_id, process_type, start_date, end_date, status, sheet=""):
    
        try:
            task = Task_CheckShape.objects.get(pk=task_id)

            ShapeCheckProcessState.objects.create(
                task=task,
                process_type = process_type,
                sheet_name=(sheet if sheet else ""),
                import_start_timestamp=start_date,
                import_end_timestamp=end_date,
                status=status
            )
            return True

        except ObjectDoesNotExist:
            logger.error(f"Task with ID {task_id} was not found: {str(e)}")
            raise
        except Exception as e:
            logger.error(f"An error occurred while importing sheet: {str(e)}")
            raise

class ShapeCalcFormulas(CalcFormulas):

    def __init__(self, *args, main_sheet=None, task_id=None, correct_values=None, columns_to_avoid=None, **kwargs):
        super().__init__(*args, **kwargs)
        self.main_sheet = main_sheet
        self.task_id = task_id
        self.correct_values = correct_values
        self.columns_to_avoid = columns_to_avoid

    def main_calc(self):
        
        # A list which will include all the columns with incorrect values (which they are not OK)
        verif_checks_results = []
        
        # Iterate through the columns in the specified range
        for col_idx in range(self.start_col, self.end_col + 1):

            start_date = timezone.now()

            col_letter = get_column_letter(col_idx)

            logger.info(f"current col : {col_letter}")

            # Get the correct value if it exists
            correct_value = self.correct_values.get(col_letter) if self.correct_values else None

            # Dictionary to set the variables of the formula
            variables = {}

            # Retrieve the formula from the first cell of the column
            formula_cell = self.sheet[f"{col_letter}{self.start_row}"]
                
            if formula_cell.data_type != 'f':  # Skip if not a formula
                continue
            
            formula = formula_cell.value
    
            # Skip a time-consuming formulas for the SHAPE checks
            if self.columns_to_avoid:
                if col_letter in self.columns_to_avoid:
                    continue
                
            # Parse and compile the formula outside the loop for better performance
            parser = formulas.Parser()
            ast = parser.ast(formula)[1]
            compiled = ast.compile()

            # Check if the formulas include the INPUT.xlsx requirement
            if re.search(r"'ANNO INPUT'", formula, re.IGNORECASE):
                # Replace the entire external reference with analysis_year
                formula = self.replace_with_year(formula)
                ast = parser.ast(formula)[1]
                compiled = ast.compile()

            # Regular expression to extract the column letters from the formula (e.g., X, L) and sheet names if exist.
            # For instance, it will catch the cell A2 but also the Fiumi!A2.
            columns_in_formula = re.findall(r'([A-Za-z_]+!)?([A-Z]{1,3})\d+', formula)
            # Remove '!' from sheet names if present
            columns_in_formula = [(match[0].rstrip('!'), match[1]) for match in columns_in_formula]
            columns_in_formula = self.exclude_range_columns(columns_in_formula, formula)
                
            # Check if the formula includes ranges e.g A:A or sheet with ranges. It catch patterns like: Fiumi_inreti!A:A
            # which is the way that Openpyxl interprets internally the actual patterns e.g $Fiumi_inreti.A:A 
            # ranges_in_formula = re.findall(r'(?:(?P<sheet>[A-Za-z_][\w]*)!)?(?P<range>\$?[A-Z]{1,3}:\$?[A-Z]{1,3})', formula)
            # this new pattern catches ranges with and without rows like B1:B2232
            ranges_in_formula = re.findall(r'(?:(?P<sheet>[A-Za-z_][\w]*)!)?(?P<range>\$?[A-Z]{1,3}:\$?[A-Z]{1,3}|\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+)', formula)   
            # Regex of row_based_ranges with catching the sheet
            row_based_ranges = re.findall(r'(?:(?P<sheet>[A-Za-z_][\w]*)!)?(?P<col1>\$?[A-Z]{1,3})(?P<row>\d+):(?P<col2>\$?[A-Z]{1,3})(?P=row)', formula)

            # Regex to capture the absolute rows like B$1 (They are used in the DB_prioritari file)
            abs_rows = re.findall(r'\b[A-Z]+\$\d+\b', formula)
            
            if ranges_in_formula:
                for match in ranges_in_formula:
                    sheet_name, col_ranges = match
                    # remove the $ from the col_ranges if exist
                    col_ranges = col_ranges.replace('$', '')
                    # Skip row-based ranges (B4:BB4, A10:C10)**

                    # Check if the reange is row_based
                    if re.match(r'^(?P<col1>[A-Z]{1,3})(?P<row>\d+):(?P<col2>[A-Z]{1,3})(?P=row)$', col_ranges):
                        continue  # Skip this range, go to next match
                    
                    if sheet_name:
                        # check if the sheet is from another file, In this case the interpreted formula will be like: [2]FIUMI!B:B
                        second_file_check = re.search(r'\[(\d+)\]', formula)
                        if second_file_check:
                            linked_number = second_file_check.group(1)
                            external_wb = self.external_link_parser(linked_number)
                           
                            col_ranges = re.sub(r'(\d+)(?=\$?\d*$)', '', col_ranges)  # Remove row number after the first column
                            formatted_variable = f"[{linked_number}]{sheet_name.upper()}!{col_ranges}"
                            variables[formatted_variable] = self.calculate_range(formula, col_ranges, sheet_name, external_wb)
                        else:
                            formatted_variable = f"{sheet_name.upper()}!{col_ranges}"  # e.g FIUMI_INRETI!A:A
                            variables[formatted_variable] = self.calculate_range(formula, col_ranges, sheet_name)
                    else:
                        variables[col_ranges] = self.calculate_range(formula, col_ranges)
                
            if row_based_ranges:

                columns_in_formula = self.exclude_cols_from_row_ranges(row_based_ranges, columns_in_formula)
                for sheet_name, col1, row, col2 in row_based_ranges:
                    if not sheet_name:
                        # the structure of row_based_ranges is: [(col1, row, col2)]
                        var = f"{col1}{row}:{col2}{row}"
                        variables[var] = self.calculate_row_based_range([col1, row, col2]) # e.g variables[B4:BB4]
                    else:
                        # the structure of row_based_ranges is: [(col1, row, col2)]
                        var = f"{sheet_name.upper()}!{col1}{row}:{col2}{row}"
                        variables[var] = self.calculate_row_based_range([col1, row, col2]) # e.g variables[sheet!B4:BB4]
            if abs_rows:
                abs_rows = self.exclude_abs_range_columns(abs_rows, ranges_in_formula)
                for i in abs_rows:
                    # the format of the abs_rows is something like "['J$1']"
                    i = i.replace("$", "")
                    variables[i] = self.sheet[f"{i}"].value

            # Iterate through each row for this column
            for row in self.sheet.iter_rows(min_row=self.start_row, max_row=self.end_row,
                                            min_col=col_idx, max_col=col_idx):
  
                cell = row[0]
                # print(row)
                # Retrieve the required values from the relevant cells
                for sheet_name, col in columns_in_formula:
                    ref_cell = f"{col}{cell.row}"
                    if not sheet_name:
                        value = self.sheet[ref_cell].value
                        # We set the text values in uppercase:
                        value = self.cell_value_parser(value)
                        # We set the key of the variables using the first row with data (4)
                        # because in that way have been compliled by Formulas. The result
                        # of course is updated with the new rows.
                        # variables[f"{col}{self.start_row}"] = value if value is not None else 0  # Default to None if 0
                        variables[f"{col}{self.start_row}"] = self.sanitize_value(value, formula)
                    else:
                        value = self.workbook[sheet_name][ref_cell].value
                        # We set the text values in uppercase:
                        value = self.cell_value_parser(value)
                        # variables[f"{sheet_name.upper()}!{col}{self.start_row}"] = value if value is not None else 0  # Default to None if 0
                        variables[f"{sheet_name.upper()}!{col}{self.start_row}"] = self.sanitize_value(value, formula)

                # Evaluate the formula with the given variables
                try:
                    calculated_result = compiled(**variables)
                except Exception as e:
                    calculated_result = f"Error: {e}"

                result = calculated_result.item() if hasattr(calculated_result, "item") else calculated_result

                # Check if the result is the correct value in case of the column checks
                if correct_value is not None:
                    if result != correct_value:
                        # store this as 1 which means that this column is not OK
                        verif_checks_results.append(col_letter)
                        cell.value = result
                        # we can use break if we want to skip the column from further caclulations
                        # break

                # print(f"Sheet: {self.sheet}, Row {cell.row} ({col_letter}{cell.row}): {result}")
                # Store the result in the target cell
                cell.value = result

            end_date = timezone.now()

            ImportStateHelper.import_column_state(self.task_id, 
                                     ProcessType.CALCULATION.value,
                                     start_date, 
                                     end_date, 
                                     TaskStatus.SUCCESS,
                                     sheet=self.main_sheet
                                    )

            logger.info(f"Column {col_letter} was calculated")
        # return the caclulated_result as single value and not as a numpy array e.g Array("OK", dtype=object)
        return (self.sheet, verif_checks_results)
    
    def replace_with_year(self, formula):
        pattern = r"\$?'ANNO INPUT'!\$?[A-Z]+\$?\d+|\$?'ANNO INPUT'![A-Z]+\d+"
        return re.sub(
                   pattern,
                   f"{self.analysis_year}",
                   formula
                   )
    
    def exclude_cols_from_row_ranges(self, row_based_ranges, columns_in_formula):
        excluded_columns = set()

        # Process the row-based ranges
        for sheet_name, start_col, row, end_col in row_based_ranges:
            # Add only the start and end columns to excluded_columns
            excluded_columns.add(start_col)
            excluded_columns.add(end_col)

        # Extract just the column names from columns_in_formula (which are the second element of the tuples)
        # Iterate over columns_in_formula and check if the column (second element) is in excluded_columns
        valid_columns = [
            (sheet_name, col) for sheet_name, col in columns_in_formula  # Keep the tuple structure intact
            if col not in excluded_columns  # Exclude columns in the row-based range
        ]

        return valid_columns

    def sanitize_value(self, value, formula):
        '''
        This method handles the empty cells like the Excel
        '''
        pattern = r'<>""'
        if value is None and re.search(pattern, formula):
            value = ""
            return value
        if value is None or (isinstance(value, float) and math.isnan(value)) or value == "":
            return 0

        return value
    

class SpecShapeCalcFormulas:
    """
    This class calculates the time-consuming formulas for the SHAPE checks
    Related issue: https://github.com/geosolutions-it/C179-DBIAIT/issues/462
    """

    def __init__(self,
                 workbook: openpyxl.workbook.Workbook,
                 sheet_name: str,
                 start_row: int,
                 end_row: int,
                 task_id = None,
                 correct_values=None
                 ):
        self.workbook = workbook
        self.sheet_name = sheet_name
        self.start_row = start_row
        self.end_row = end_row
        self.task_id = task_id
        self.correct_values = correct_values

    def simple_countif(self, col):
        '''
        This function calculates the formula below:
        +IF(COUNTIF(D:D,D5)=1,0,1)
        '''
        start_date = timezone.now()
        # Get the correct value if it exists
        correct_value = self.correct_values.get(col) if self.correct_values else None
        incorrect_value = False
        # sheet definition
        ws = self.workbook[self.sheet_name]

        data = ws.values

        # Convert to DataFrame
        df = pd.DataFrame(data)

        start_idx = self.start_row - 1
        end_idx = self.end_row
        
        # the index for column D in pandas is 3
        column_d = df.iloc[start_idx:end_idx, 3]

        value_counts_all = column_d.value_counts()

        calculated_values = column_d.map(lambda x: 0 if value_counts_all[x] == 1 else 1)

        # check if the calculated_values includes incorrect values
        if (calculated_values != correct_value).any():
            incorrect_value = True

        end_date = timezone.now()
        ImportStateHelper.import_column_state(self.task_id, 
                                     ProcessType.CALCULATION.value,
                                     start_date, 
                                     end_date, 
                                     TaskStatus.SUCCESS,
                                     sheet=self.sheet_name
                                    )
        
        return calculated_values, incorrect_value
    
    def countif_with_sheet(self, col):
        '''
        This method calculates the formulas below:
        =+IF(N5="DISTRIBUZIONE",IF(COUNTIF(Distrib_tronchi!B:B,D5)>0,0,1),IF(COUNTIF(Addut_tronchi!B:B,D5)>0,0,1))
        =+IF(O5="FOGNATURA",IF(COUNTIF($Fognat_tronchi.$B:$B,D5)>0,0,1),IF(COUNTIF($Collett_tronchi.$B:$B,D5)>0,0,1))
        '''
        start_date = timezone.now()
        # Get the correct value if it exists
        correct_value = self.correct_values.get(col) if self.correct_values else None
        incorrect_value = False
        
        if self.sheet_name == "SHP_Acquedotto":
            col_var = "N"
            col_value = "DISTRIBUZIONE"
            sheet1 = "Distrib_tronchi"
            sheet2 = "Addut_tronchi"
        elif self.sheet_name == "SHP_Fognatura":
            col_var = "O"
            col_value = "FOGNATURA"
            sheet1 = "Fognat_tronchi"
            sheet2 = "Collett_tronchi"
        else:
            logger.info("The sheet was not found during the specialized formulas calculations")
       
       # Defind the column indices for the col_var
        col_var_idx = column_index_from_string(col_var)
        col_var_idx -= 1
    
       # Get the correct value if it exists
        correct_value = self.correct_values.get(col) if self.correct_values else None
        incorrect_value = False
        # dataframe of the main sheet
        ws = self.workbook[self.sheet_name]
        data = ws.values

        # Convert to DataFrame
        df = pd.DataFrame(data).dropna(how='all')  # Remove fully empty rows

        # dataframes of the relative sheets
        ws_sheet1 = self.workbook[sheet1]
        sheet1_data = ws_sheet1.values

        # Convert to DataFrame
        df_sheet1 = pd.DataFrame(sheet1_data).dropna(how='all')

        ws_sheet2 = self.workbook[sheet2]
        sheet2_data = ws_sheet2.values

        # Convert to DataFrame
        df_sheet2 = pd.DataFrame(sheet2_data).dropna(how='all')

        start_idx = self.start_row - 1
        end_idx = self.end_row

        df_subset = df.iloc[start_idx:end_idx, :].copy()

        df_subset = df_subset.rename(columns={0: "A", 
                                              col_var_idx: col_var, 
                                              3: "D"})  # Column N or O (Condition), Column D (Value to check)

        # In other sheets beyond the main sheet, the start row is the row 4
        sheet1_set = set(df_sheet1.iloc[3:, 1])  # Column B of Distrib_tronchi
        sheet2_set = set(df_sheet2.iloc[3:, 1])  # Column B of Addut_tronchi

        # Apply the formula logic using .map()
        def apply_formula(row):
            if row[col_var] == col_value:
                value_exists = row["D"] in sheet1_set  # Check if D5 exists in Distrib_tronchi B column
            else:
                value_exists = row["D"] in sheet2_set  # Check if D5 exists in Addut_tronchi B column

            return 0 if value_exists else 1

        df_subset["Calculated_Value"] = df_subset.apply(apply_formula, axis=1)

        # Check if the calculated values match the expected correct values
        if correct_value is not None:
            if (df_subset["Calculated_Value"] != correct_value).any():
                incorrect_value = True
        
        end_date = timezone.now()
        ImportStateHelper.import_column_state(self.task_id, 
                                     ProcessType.CALCULATION.value,
                                     start_date, 
                                     end_date, 
                                     TaskStatus.SUCCESS,
                                     sheet=self.sheet_name
                                    )
        
        return (df_subset["Calculated_Value"], incorrect_value)


    def lookup_with_sheet(self, col):
        '''
        This method calculates the formulas below:
        =+IF(N5="DISTRIBUZIONE",IF(VLOOKUP(A5,$Distribuzioni.$B:$AD,29,FALSE())<3,0,1),IF(VLOOKUP(A5,$Adduttrici.$B:$Z,25,FALSE())<3,0,1))
        =+IF(O5="FOGNATURA",IF(VLOOKUP(A5,$Fognature.B:T,19,FALSE())<3,0,1),IF(VLOOKUP(A5,$Collettori.B:N,13,FALSE())<3,0,1))
        '''
        start_date = timezone.now()
        # Get the correct value if it exists
        correct_value = self.correct_values.get(col) if self.correct_values else None
        incorrect_value = False
        
        if self.sheet_name == "SHP_Acquedotto":
            col_var = "N"
            col_value = "DISTRIBUZIONE"
            sheet1 = "Distribuzioni"
            sheet2 = "Adduttrici"
            start_range_col1 = "B"
            end_range_col1 = "AD"
            start_range_col2 = "B"
            end_range_col2 = "Z"

        elif self.sheet_name == "SHP_Fognatura":
            col_var = "O"
            col_value = "FOGNATURA"
            sheet1 = "Fognature"
            sheet2 = "Collettori"
            start_range_col1 = "B"
            end_range_col1 = "T"
            start_range_col2 = "B"
            end_range_col2 = "N"
        else:
            logger.info("The sheet was not found during the specialized formulas calculations")
       
       # Define the column indices for the col_var
        col_var_idx = column_index_from_string(col_var)
        col_var_idx -= 1

        # Define end col range indices
        start_range_col1_idx = column_index_from_string(start_range_col1)
        start_range_col1_idx -= 1
        start_range_col2_idx = column_index_from_string(start_range_col2)
        start_range_col2_idx -= 1
        end_range_col1_idx = column_index_from_string(end_range_col1)
        end_range_col1_idx -= 1
        end_range_col2_idx = column_index_from_string(end_range_col2)
        end_range_col2_idx -= 1
    
       # Get the correct value if it exists
        correct_value = self.correct_values.get(col) if self.correct_values else None
        incorrect_value = False

        # dataframe of the main sheet
        ws = self.workbook[self.sheet_name]
        data = ws.values

        # Convert to DataFrame
        df = pd.DataFrame(data).dropna(how='all')  # Remove fully empty rows
        
        start_idx = self.start_row - 1
        end_idx = self.end_row

        # Setup the dataframe for the related sheets
        ws_sheet1 = self.workbook[sheet1]
        sheet1_data = ws_sheet1.values
        # Convert to DataFrame
        df_sheet1 = pd.DataFrame(sheet1_data).dropna(how='all')

        ws_sheet2 = self.workbook[sheet2]
        sheet2_data = ws_sheet2.values
        # Convert to DataFrame
        df_sheet2 = pd.DataFrame(sheet2_data).dropna(how='all')

        df_subset = df.iloc[start_idx:end_idx, :].copy()

        df_subset = df_subset.rename(columns={0: "A", 
                                              col_var_idx: col_var})  # Column A (ID) & Column N (Condition)

        sheet1_lookup = df_sheet1.set_index(start_range_col1_idx)[end_range_col1_idx].to_dict() # Column B -> Column AD or T
        sheet2_lookup = df_sheet2.set_index(start_range_col2_idx)[end_range_col2_idx].to_dict()  # Column B -> Column Z or N

        def apply_formula(row):
            if row[col_var] == col_value:
                value = sheet1_lookup.get(row["A"], 0)  # Default to 0
            else:
                value = sheet2_lookup.get(row["A"], 0)  # Default to 0

            # Handle NoneType explicitly
            if value is None:
                value = float("inf")  # Default to a large number

            # Convert to float if it's a valid number
            try:
                value = float(value)
            except (ValueError, TypeError):  
                value = float("inf")  # If conversion fails, return 0

            return 0 if value < 3 else 1

        df_subset["Calculated_Value"] = df_subset.apply(apply_formula, axis=1)

        # Check if the calculated values match the expected correct values
        if correct_value is not None:
            if (df_subset["Calculated_Value"] != correct_value).any():
                incorrect_value = True
        
        end_date = timezone.now()
        ImportStateHelper.import_column_state(self.task_id, 
                                     ProcessType.CALCULATION.value,
                                     start_date, 
                                     end_date, 
                                     TaskStatus.SUCCESS,
                                     sheet=self.sheet_name
                                    )
        
        return (df_subset["Calculated_Value"], incorrect_value)




