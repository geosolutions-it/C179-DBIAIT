import os
import json
import pathlib
import pandas as pd
from dbfread import DBF
import gc
from collections import defaultdict

from openpyxl.formula.translate import Translator
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import numbers

from django.utils import timezone
from django.db.models import ObjectDoesNotExist
from django.conf import settings

from app.dbi_checks.models import TaskStatus, ProcessType
from app.dbi_checks.utils import YearHandler
from app.dbi_checks.tasks.checks_definitions.base_calc import BaseCalc

from app.shape_checks.tasks.checks_definitions.shape_formulas_calc import (
    ShapeCalcFormulas, 
    SpecShapeCalcFormulas
    )
from app.shape_checks.models import Task_CheckShape, ShapeCheckProcessState


import logging

logger = logging.getLogger(__name__)


class ShapeCalc(BaseCalc):
    
    def __init__(
        self, 
        orm_task, # Task_CheckShape
        imported_file: str,
        imported_dbf_file: str,
        sheet_for_dbf: str,
        seed: str,
        config: str,
        formulas_config: str,
        export_dir: pathlib.Path,
        file_year_required: bool = False,
        task_progress: int = 0,
        log_workbook = None,
        summary_data = None,
    ):
        super().__init__(orm_task,
                         imported_file,
                         seed,
                         config,
                         formulas_config,
                         export_dir,
                         file_year_required,
                         task_progress,
                         log_workbook,
                         summary_data,
                         )
        self.imported_dbf_file = imported_dbf_file
        self.sheet_for_dbf = sheet_for_dbf
    
    def drag_formulas(self, seed_wb):
        
        # Before dragging the formulas we have to copy
        # the DBF content to the specialized sheet
        dbf_copy_result = self.copy_from_dbf(seed_wb)

        if dbf_copy_result:
            super().drag_formulas(seed_wb)
    
    def copy_from_dbf(self, seed_wb):
        try:
            start_date = timezone.now()
            # Read DBF file into a DataFrame
            dbf_table = DBF(self.imported_dbf_file, load=True)
            df = pd.DataFrame(iter(dbf_table))

            with open(settings.DBF_TO_SHEET, "r") as file:
                dbf_to_sheet_config = json.load(file)

            self.orm_task.progress += self.task_progress
        except (FileNotFoundError, json.JSONDecodeError) as e:
            logger.error(f"Error reading DBF file: {e}")
            return False
        
        try:
            # Get the sheet name where we will copy the DBF data
            sheet_mapping = dbf_to_sheet_config.get(self.sheet_for_dbf, {})
        
            start_row = sheet_mapping.get('start_row', 5)    
            start_col_letter = sheet_mapping.get('start_col', "A")
            start_col = column_index_from_string(start_col_letter)
            
            total_rows = len(df)
            total_cols = len(df.columns)

            # Load existing Excel sheet
            sheet_obj = seed_wb[self.sheet_for_dbf]

            logger.info(f"Start copying from DBF to Excel")
        
            # Use iter_rows() to access cells and assign values
            for row, df_row in zip(
                sheet_obj.iter_rows(min_row=start_row, max_row=start_row + total_rows - 1,
                                    min_col=start_col, max_col=start_col + total_cols - 1),
                df.itertuples(index=False, name=None)
            ):
                for cell, value in zip(row, df_row):
                    cell.value = value  # Assign value to each cell

            end_date = timezone.now()
            
            self.import_process_state(self.orm_task.id, 
                                          ProcessType.COPY.value,
                                          os.path.basename(self.imported_dbf_file), 
                                          start_date, 
                                          end_date, 
                                          TaskStatus.SUCCESS,
                                          sheet=self.sheet_for_dbf
                                          )

            logger.info(f"Copied {total_rows} rows from DBF to Excel")

            self.orm_task.progress += self.task_progress

            return True
        
        except (KeyError, Exception) as e:
            end_date = timezone.now()
            self.import_process_state(self.orm_task.id, 
                                          ProcessType.COPY.value,
                                          os.path.basename(self.imported_dbf_file), 
                                          start_date, 
                                          end_date, 
                                          TaskStatus.FAILED,
                                          sheet=self.sheet_for_dbf
                                          )
            logger.error(f"Error copying data to Excel: {e}")
            return False

    def log_file_manager(self, seed_wb, seed_name):
        
        ## Configuration setup
        # prepare the logs workbook
        # Remove default sheet if it exists
        if "Sheet" in self.log_workbook.sheetnames and len(self.log_workbook.sheetnames) == 1:
            del self.log_workbook["Sheet"]
        
        sheet_name = "Logs"
        if sheet_name not in self.log_workbook.sheetnames:
            log_sheet = self.log_workbook.create_sheet(sheet_name)
            log_sheet.append(["File", 
                              "Foglio",
                              "Codice opera",
                              "Colonna check", 
                              "Descrizione", 
                              "Valore colonna check", 
                              "Valore colonna 1", 
                              "Valore colonna 2",
                              "Valore colonna 3",
                              "Valore colonna 4"

                              ])
        else:
            log_sheet = self.log_workbook[sheet_name]
        
        # Set the style in the log file
        self.set_logfile_style(log_sheet)
                
        # set the configs
        analysis_year = YearHandler(self.imported_file).get_year()

        # Get the seed_file name in order to retrieve it
        # from the log_mapping.json

        seed_filename = pathlib.Path(self.seed).stem
        seed_key = seed_filename.upper()
                
        with open(settings.LOG_MAPPING, "r") as file:
            log_mapping = json.load(file)
                
        verif_checks_config = log_mapping.get(seed_key, {})

        formulas_config = self.load_formulas_conf(seed_key)

        # Load the specialized (time-consuming) formulas
        with open(settings.SPEC_SHAPE_FORMULAS, "r") as file:
            spec_shape_formulas = json.load(file)
        
        ## Calculate the formulas of the checks for each sheet
        for sheet_name, f_location in formulas_config.items():

            pd_sheet = None
            start_date = timezone.now()
            
            if sheet_name in seed_wb.sheetnames:
                sheet = seed_wb[sheet_name]

                # Get column indexes
                start_col_index = column_index_from_string(f_location["start_col"])
                end_col_index = column_index_from_string(f_location["end_col"])
                start_row = f_location["start_row"]
                # Re-definition of the last row because the copied file is processed
                # without saving yet. We don't want to re-load it for time reasons
                end_row = self.get_end_row(f_location, sheet_name, sheet)

                ## setup the config for each sheet
                sheet_checks = verif_checks_config.get(sheet_name, None)

                calculator = self.get_calculator()

                # Map the check columns with the corresponging correct values
                correct_values = {item["colonna_check"]: item["valore"] for item in sheet_checks}
                if sheet_name in {"Controllo dati aggregati", "Controlli aggregati"}:
                    
                    # we call the calculator two times because in these sheets
                    # there are different formulas in each row
                    for row in [start_row, end_row]:
                        sheet_with_calc_values, verif_checks_results = calculator(
                            workbook=seed_wb, 
                            sheet=seed_wb[sheet_name],
                            main_sheet=sheet_name,
                            start_row=row, 
                            end_row=row,  # Keep start and end row the same for individual calculations
                            start_col=start_col_index,
                            end_col=end_col_index,
                            analysis_year=analysis_year,
                            external_wb_path=self.export_dir,
                            task_id=self.orm_task.id,
                        ).main_calc()
                else:
                    # a list to store the specialized column that the first calculation
                    # has to avoid
                    columns_to_avoid = []
                    # Calculate the specialized (time-consuming) formulas
                    spec_shape_formulas_config = spec_shape_formulas[sheet_name]
                    
                    for f in spec_shape_formulas_config.get("spec_formulas", {}):
                        col = f["col"]
                        columns_to_avoid.append(col)
                    # Calculate the main column checks
                    sheet_with_calc_values, verif_checks_results = calculator(workbook=seed_wb, 
                                                sheet=seed_wb[sheet_name],
                                                main_sheet=sheet_name,
                                                start_row=start_row,
                                                # temp end row for testing
                                                end_row = end_row,
                                                start_col = start_col_index,
                                                end_col = end_col_index,
                                                analysis_year=analysis_year,
                                                external_wb_path=self.export_dir,
                                                task_id=self.orm_task.id,
                                                correct_values = correct_values,
                                                columns_to_avoid = columns_to_avoid
                                                ).main_calc()

                    # Initialization of the SpecShapeClass
                    # Related issue: https://github.com/geosolutions-it/C179-DBIAIT/issues/462
                    spec_shape_calc_formulas_instance = SpecShapeCalcFormulas(
                            seed_wb,
                            sheet_name,
                            start_row,
                            end_row,
                            self.orm_task.id,
                            correct_values
                        )
                    # Dict with with the calculated columns
                    calc_spec_columns = {}
                    for f in spec_shape_formulas_config.get("spec_formulas", {}):
                        col = f["col"]
                        method_name = f['method_name']
                        # Dynamically call the method based on the JSON mapping
                        if hasattr(spec_shape_calc_formulas_instance, method_name):
                            method = getattr(spec_shape_calc_formulas_instance, method_name)
                            logger.info(f"Processing {col} using {method}")
                            calc_spec_columns[col], incorrect_value = method(col)
                            # Check if the result includes incorrect values
                            if incorrect_value:
                                verif_checks_results.append(col)

                        else:
                            logger.info(f"Method {method_name} not found in class.")
                
                for check in sheet_checks:
        
                    # Get the verification check cell to check if it is OK or not
                    verif_check = check.get("verif_check", {})

                    column_check = check.get("colonna_check", None)
                    column_rel = check.get("colonna_rel", None)
                    desc = check.get("descrizione", None)
                    criterion = check.get("valore", 0)
                    
                    if verif_check is None:
                        continue

                    if pd_sheet is None:
                        # Read all data from the worksheet
                        data = list(sheet.iter_rows(min_row=1, 
                                                            max_row=self.get_last_data_row(sheet_with_calc_values), 
                                                            values_only=True))  # Read all rows as tuples

                        # Convert to DataFrame
                        pd_sheet = pd.DataFrame(data)
                        logger.info(f"The data frame of the sheet {sheet_name} was created")

                        # Extend the pd_sheet in order to include the specialized columns of the SHAPE checks
                        for key, value in calc_spec_columns.items():
                            col_idx = self.parse_col_for_pd(key)
                            row_idx = start_row - 1
                            pd_sheet.loc[row_idx:, col_idx] = value

                    # Get rows where column_check is NOT 0
                    start_idx = start_row - 1
                    if sheet_name in {"Controllo dati aggregati", "Controlli aggregati"}:
                            
                        # convert column names for pandas
                        column_check_idx = self.parse_col_for_pd(column_check)
                        # column_rel_idx = self.parse_col_for_pd(column_rel)
                        
                        verif_check_col = verif_check["col"]
                        verif_check_col_index = column_index_from_string(verif_check_col)
                        verif_check_row = verif_check["row"]
                        
                        # caclulate the formula in the verification check
                        sheet_with_verif_values,  verif_checks_results = calculator(workbook=seed_wb, 
                                                sheet=sheet_with_calc_values,
                                                main_sheet=sheet_name,
                                                start_row=verif_check_row, 
                                                end_row = verif_check_row,
                                                start_col = verif_check_col_index,
                                                end_col = verif_check_col_index,
                                                analysis_year=analysis_year,
                                                external_wb_path=self.export_dir,
                                                task_id=self.orm_task.id,
                                                ).main_calc()
                        
                        # retrieve the calculated verif check value
                        verif_check_value = sheet_with_verif_values[f"{verif_check_col}{verif_check_row}"].value
                        if verif_check_value != "OK":
                            logger.info(f"The check of the cell {verif_check_col}{verif_check_row} is not OK")

                            end_idx = end_row - 1

                            rows_to_check = [start_idx, end_idx]  # Only check start and end rows
                            filtered_rows = pd_sheet.iloc[rows_to_check]
                            filtered_rows = filtered_rows[filtered_rows[column_check_idx] != criterion]

                            self.verbose_log_file(column_check_idx,
                                                  sheet_name,
                                                  column_rel,
                                                  desc,
                                                  seed_key,
                                                  column_check,
                                                  log_sheet,
                                                  filtered_rows
                                                  )
                        else:
                            logger.info(f"The column check {column_check} is OK")

                    else:
                        if column_check in verif_checks_results:
                        
                            logger.info(f"The check column {column_check} is NOT OK")
                            
                            # convert column names for pandas
                            column_check_idx = self.parse_col_for_pd(column_check)
                            # column_rel_idx = self.parse_col_for_pd(column_rel)
                            
                            # Ensure column_rel is always a list (or empty if None)
                            if column_rel is None:
                                column_rel = []


                            filtered_rows = pd_sheet.iloc[start_idx:][
                                    (pd_sheet.iloc[start_idx:, column_check_idx] != criterion) & 
                                    (pd_sheet.iloc[start_idx:, column_check_idx].notna()) & 
                                    pd_sheet.iloc[start_idx:, column_check_idx].apply(lambda x: isinstance(x, (int, float)))  # Ensure it's numeric
                                ]
                            logger.info("filtered_rows from pandas where created")
                        
                            self.verbose_log_file(
                                column_check_idx,
                                sheet_name,
                                column_rel,
                                desc,
                                seed_key,
                                column_check,
                                log_sheet,
                                filtered_rows
                            )
                        else:
                            logger.info(f"The column check {column_check} is OK")
                
                end_date = timezone.now()

                self.import_process_state(self.orm_task.id, 
                                        ProcessType.LOG.value,
                                        os.path.basename(self.imported_file), 
                                        start_date, 
                                        end_date, 
                                        TaskStatus.SUCCESS,
                                        sheet=sheet_name
                                        )
            
            # Remove DataFrame from memory and trigger garbage collection
            del pd_sheet
            gc.collect()
        
        self.task_progress = self.task_progress + 20
    
    def import_process_state(self, task_id, process_type, file_name, start_date, end_date, status, sheet=""):
    
        try:
            task = Task_CheckShape.objects.get(pk=task_id)

            ShapeCheckProcessState.objects.create(
                task=task,
                process_type = process_type,
                sheet_name=(sheet.lower() if sheet else ""),
                file_name=file_name,
                import_start_timestamp=start_date,
                import_end_timestamp=end_date,
                status=status
            )

            logger.info(f"Successfully imported sheet: {sheet} for task ID {task_id}")
            return True

        except ObjectDoesNotExist:
            logger.error(f"Task with ID {task_id} was not found: {str(e)}")
            raise
        except Exception as e:
            logger.error(f"An error occurred while importing sheet: {str(e)}")
            raise
    
    def year_to_file(self, seed_wb):
        # Write the year to the resulted file
        defined_year = YearHandler(self.imported_file).get_year()
        anno_sheet = seed_wb["ANNO INPUT"]
        anno_sheet['A1'] = defined_year
        logger.info(f"The year {defined_year} was copied to the ANNO INPUT sheet")

    def get_the_unique_code(self, sheet_name, row):
            
        if sheet_name not in {"Controllo dati aggregati", "Controlli aggregati"}:
            # retrieve the column D which includes the unique code (COD_TRATTO) of each record
            unique_code_idx = self.parse_col_for_pd('D')
            unique_code = row[unique_code_idx]
            return unique_code
        else:
            return None
        
    def get_end_row(self, f_location, sheet_name, sheet):
        if sheet_name in {"Controllo dati aggregati", "Controlli aggregati"}:
            end_row = f_location.get("end_row", 3)
        else:
            end_row = self.get_last_data_row(sheet)
        return end_row
    
    def load_formulas_conf(self, seed_key):
        # Open the json file with the verif shape formulas
        with open(settings.SHAPE_VERIF_FORMULAS, "r") as file:
            shape_verif_formulas = json.load(file)
        formulas_config = shape_verif_formulas.get(seed_key, {})
        return formulas_config
    
    def get_calculator(self):
        return ShapeCalcFormulas
    
    def verbose_log_file(self, column_check_idx, 
                         sheet_name, 
                         column_rel, 
                         desc, 
                         seed_key, 
                         column_check, 
                         log_sheet, 
                         filtered_rows,
                         ):
        # Iterate through the filtered rows and retrieve the necessary information
        for index, row in filtered_rows.iterrows():
            incorrect_value = row[column_check_idx]  # Value of the cell in `colonna_check`
                            
            # get the unique code (Codice opera)
            unique_code = self.get_the_unique_code(sheet_name, row)
                            
            # Handle the colonna_rel values:
            # Initialize a dictionary to store values from each related column
            related_values_dict = {}
            for rel_col in column_rel:  # Iterate through all columns in column_rel list
                # Ensure to get the value of each related column (if it exists)
                if rel_col:
                    related_value = row[self.parse_col_for_pd(rel_col)] if rel_col else None
                    related_values_dict[rel_col] = related_value

            # Replace placeholders in the description with the corresponding values from related_values_dict
            # we re-define the updated_desc to be equal with the initial desc
            updated_desc = desc
            for key, value in related_values_dict.items():
                placeholder = f"{{{key}}}"  # e.g., {AO}
                if placeholder in updated_desc:
                    updated_desc = updated_desc.replace(placeholder, str(value))

            # Create a list for the related values, using the dictionary's get method to handle missing keys
            related_values = [related_values_dict.get(key, None) for key in list(related_values_dict.keys())]

            # Append the row to the log sheet
            log_sheet.append([seed_key, sheet_name, unique_code, column_check, updated_desc, incorrect_value] + related_values)

            # Track summary entry
            self.summary_data[(seed_key, sheet_name, column_check)] += 1

    